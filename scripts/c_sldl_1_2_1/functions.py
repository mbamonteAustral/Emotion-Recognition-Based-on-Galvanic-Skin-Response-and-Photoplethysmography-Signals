# ----- PRE-PROCESSING STAGE (GSR + PGG) ----------

import time
import numpy as np
import matplotlib.pyplot as plt
import random
import neurokit2 as nk # Load the NeuroKit package
import heartpy as hp  #a HR analysis toolkit designed for PPG | https://pypi.org/project/heartpy/
import tensorflow as tf
import pandas as pd
from scipy import signal
from scipy.signal import medfilt, find_peaks, lfilter
from scipy import signal
from scipy.fft import fft, fftfreq
from collections import Counter
from neurokit2 import signal_detrend
from tensorflow import keras
from tensorflow.keras import layers, regularizers, initializers
from openpyxl import load_workbook
from math import isnan 
from sys import platform
from definitions import *

from sklearn.model_selection import StratifiedShuffleSplit, KFold, RepeatedStratifiedKFold, RepeatedKFold, cross_val_score, cross_validate
from sklearn.svm import SVC
from sklearn.neighbors import KNeighborsClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.ensemble import GradientBoostingClassifier
from sklearn.dummy import DummyClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import StandardScaler, MinMaxScaler
from sklearn.pipeline import Pipeline
from numpy import mean, std
from imblearn.metrics import geometric_mean_score
from sklearn.metrics import make_scorer, balanced_accuracy_score, precision_score, classification_report, accuracy_score, f1_score, fbeta_score, cohen_kappa_score, recall_score, confusion_matrix

RepeatedStratifiedKFold

# My own modules (current folder modules)
from definitions import *




# -----------------------------------------------------------------------
# CASE processing functions


def map_video_to_tag (video_id):
    """
    map Video-ID to stimulus tag
    Video-ID  >= 10 will be assigned stimulus tag = 1, corresponding to the blue screen/initial screen.
    Video-ID < 10 will be assigned stimulus tag = 2, corresponding to a true stimulus.
    """
    if video_id >= 10:
        stimulus_tag = NEUTRAL_SCREEN
    elif video_id < 10:
        stimulus_tag = STIMULUS_SCREEN
    return stimulus_tag




def scale_tag (tag,min_value,max_value):
    """
    Scales a two valued vector (Low/high) to the min / max value of a signal.
    It's useful to visualize the stimulus timestamps in the physiological signal altogether 

    """ 
    if tag == NEUTRAL_SCREEN:
        scaled_tag = min_value
    elif tag == STIMULUS_SCREEN:
        scaled_tag = max_value
    return scaled_tag



def baseline_mean_centered_normalization (data, show_plot = 0):
    """
    baseline Mean-centered normalization. 
    Base on Zitouni et al (2023). https://ieeexplore.ieee.org/document/9965601/

    Calculate the mean of the baseline (the last 60 seconds of the starting video; CASE has a 101-second baseline) 
    and subtract it from all the time-series.

    data Sn = S - mean(Sr), where Sr the last 60 sec of the baseline (start video)
    In this case, Sn = gsr_mc or Sr = ppg_mc

    Expects: data: pandas DataFrame. 
    Returns: pandas Series: Baseline mean-centered normalization of PPG and GSR. 
    """

    # START_VIDEO_DURATION = 101500   # ms 
    # BASELINE_WINDOW = 60000 # ms 

    ppg_baseline_mean = data['bvp'][START_VIDEO_DURATION - BASELINE_WINDOW : START_VIDEO_DURATION].mean()
    gsr_baseline_mean = data['gsr'][START_VIDEO_DURATION - BASELINE_WINDOW : START_VIDEO_DURATION].mean()

    ppg_mc = data['bvp'] - ppg_baseline_mean  # mean centered based normalization
    gsr_mc = data['gsr'] - gsr_baseline_mean # mean centered based normalization


    if show_plot:
        plt.plot(data['bvp'])
        t = np.linspace(0, data.shape[0], data.shape[0], endpoint=False)
        plt.plot(t,np.ones(data.shape[0])*ppg_baseline_mean,'r')
        plt.plot(ppg_mc, label="with baseline removed")
        plt.legend()
        plt.show(block=True)   #prevent GUI event loop to take control and wait all data to be display before releasing control to the main code loop.

        plt.plot(data['gsr'])
        t = np.linspace(0, data.shape[0], data.shape[0], endpoint=False)
        plt.plot(t,np.ones(data.shape[0])*gsr_baseline_mean,'r')
        plt.plot(gsr_mc, label="with baseline removed")
        plt.legend()
        plt.show(block=True)   #prevent GUI event loop to take control and wait all data to be display before releasing control to the main code loop.

    return gsr_mc, ppg_mc



def plot_signals_with_stimulus_timestamps(ppg_mc,gsr_mc, stimulus_tags):
    """
    Plot GSR + PPG signal with stimulus timestamps.
    gsr_mc: pandas Series with GSR signal
    ppg_mc: pandas Series with PPG signal
    stimulus_tags: pandas Series with annotations (labels)
    """

    plt.ion()
    ax1=plt.subplot(211)
    plt.title("PPG")
    plt.ylabel("[%]")
    plt.grid()
    ax2=plt.subplot(212)
    plt.title("GSR")
    plt.ylabel("Ohms")
    plt.grid()
    plt.xlabel("Time [min]")

    gsr_min_value = min(gsr_mc)
    gsr_max_value = max(gsr_mc)
    gsr_scaled_tag = [scale_tag(i,gsr_min_value,gsr_max_value) for i in stimulus_tags] 

    min_value = min(ppg_mc)
    max_value = max(ppg_mc)
    ppg_scaled_tag = [scale_tag(i,min_value,max_value) for i in stimulus_tags] 


    ax1.plot(time_vector,ppg_mc, label='PPG')
    ax1.plot(time_vector,ppg_scaled_tag, label='Low/High = Neutral/stimulus screen')
    plt.grid()
    ax2.plot(time_vector,gsr_mc, label='GSR')
    ax2.plot(time_vector,gsr_scaled_tag, label='Low/High = Neutral/stimulus screen')
    plt.grid()
    plt.show(block=True)   #prevent GUI event loop to take control and wait all data to be display before releasing control to the main code loop.



def filter_gsr(gsr_signal, show_plot=0):
    """
    Extract the cleaned GSR signal from the baseline mean-centered normalized signal
    GSR filter parameters:
    1) digital IIR Butterworth filter with SOS method
    2) 4th order Butterworth low-pass filter with cut-off freq = 3 Hz
    
    Returns: 
    - gsr_signals with filtered signal, tonic and phasic signals, among other. 
    - gsr_info: "info (dict): A dictionary containing the information of each SCR peak (see eda_findpeaks()), as well as the signals’ sampling rate."
    See https://neuropsychology.github.io/NeuroKit/functions/eda.html
    
    """ 

    # Process the raw EDA signal
    gsr_signals, gsr_info = nk.eda_process(gsr_signal, sampling_rate=SAMPLING_RATE, method="neurokit")

    if show_plot:
        # Plot EDA signal
        nk.eda_plot(gsr_signals, sampling_rate=SAMPLING_RATE)

    return gsr_signals, gsr_info


def filter_ppg(ppg_signal, show_plot=0):
    """
    Extract the cleaned PPG signal from the baseline mean-centered normalized signal
    butterworth Bandpass filter parameters: 
    lowcut=0.5, highcut=8, order=3
    Returns: ppg_signals with filtered signal, tonic and phasic signals, among other. 
    See https://neuropsychology.github.io/NeuroKit/functions/eda.html
    
    """ 

    # Process the raw PPG signal
    ppg_signals, ppg_info = nk.ppg_process(ppg_signal, sampling_rate=SAMPLING_RATE)


    if show_plot:
        # Plot PPG signal
        nk.ppg_plot(ppg_signals, sampling_rate=SAMPLING_RATE)
        plt.show(block=True)   #prevent GUI event loop to take control and wait all data to be display before releasing control to the main code loop.

    return ppg_signals


def compare_onsets(gsr_signals, gsr_info ,stimulus_tag):
    """
    Compare onsets (events) with stimulus timestamps. 
    Onsets are shown with black dotted vertical lines, 
    while stimulus timestamps are shown in red lines.
    """

    # Extract clean EDA and SCR features
    cleaned = gsr_signals["EDA_Clean"]
    features = [gsr_info["SCR_Onsets"]]

    # Visualize SCR features in cleaned EDA signal
    nk.events_plot(features, cleaned, color=['black'])

    gsr_min_value = min(cleaned)
    gsr_max_value = max(cleaned)
    gsr_scaled_tag = [scale_tag(i,gsr_min_value,gsr_max_value) for i in data['tag']] 

    plt.plot(time_vector,gsr_scaled_tag,'r')
    plt.show(block=True)



def segment_time_series(time_series, window_size, overlap=False, stride=1):
    """
    Perform windowing / segmentation on a time-series.
    The window size is set to window_size.
    If overlap is set to true, the sliding window step is equal to the stride; otherwise, the step is set to window_size.
    In the former case, the percentage of overlap is calculated as [(stride / window size) * 100].
    If the length of the time_series is not a multiple of the step, the last segment will be lost and not processed.
    
    """
    segments = []
    step = stride if overlap else window_size
    # index = []
    for i in range(0, len(time_series) - window_size + 1, step):
        segment = time_series[i:i + window_size]
        segments.append(segment)
        # index.append(i)

    # return segments, index
    return segments


def test_segment_time_series(time_series_data, window_size,step,overlap, segment_indexes):
    """
    Test_windowing
    It display figures
    time_series_data: any time-series
    window_size: window size in ms.
    step: stride of sliding window, in ms.
    overlap: True / false
    segment_indexex: list with indexes to check overlap (e.g, segment_indexes = [0,1,2,3,814,815])
    """
    
    segments = segment_time_series(time_series_data, window_size, overlap,step)

    # Verification of correct overlapping
    
    stride = step if overlap else window_size
    max_number_of_segments = int(len(time_series_data)/stride)

    if max(segment_indexes) < max_number_of_segments:  
        for i in segment_indexes:
            plt.figure()
            plt.plot(time_series_data)
            plt.plot(segments[i][:])
            plt.title("Check overlapping")
        plt.show(block=True)
    else: 
        print("ERROR: maximum index of segment_indexes is greater than max_number_of_segments")


def upsample_annotations(annotation, _sampling_rate=20, _desired_sampling_rate = 1000, _method = "numpy", show_plot=1):
    """
    Upsample labels / annotations time_series (to have the same sample rate as biosignals)
    Verification of correct annotations upsampling

    annotation: valence or arousal annotation
    sampling_rate and desired_sampling_rate in Hz. 
    method = "numpy" (default). Others: ["interpolation", "FFT", "poly", "pandas"]

    returns: upsampled annotation time-serie.
    """

    annotation_upsampled = nk.signal_resample(annotation, sampling_rate=_sampling_rate, desired_sampling_rate=_desired_sampling_rate, method=_method)

    if show_plot:
        # numpy preserves original waveform:
        plt.figure()
        plt.plot(annotation_upsampled,label='numpy')
        plt.plot(annotation, label='original')
        plt.legend()

        plt.figure()
        plt.plot(annotation_upsampled,label='numpy')
        plt.legend()
        plt.figure()
        plt.plot(annotation, label='original')
        plt.legend()
        plt.show(block=True)

    return annotation_upsampled


def annotation_tag_segmentation(valence_upsampled, arousal_upsampled, time_vector, data_tag, window_size, overlap,step, show_plot=1):
    """
    windowing / segmentation on the annotations, time-vector and video tag. 
    returns: valence, arousal,time-vector and tag segments [val_segments,aro_segments, time_vector_segments, tag_segments] with the provided window_size, overlap,step
    """
    

    valence_ups = pd.Series(valence_upsampled)  #annotation must be a pandas series, because segment_time_series expects data in this format
    val_segments = segment_time_series(valence_ups, window_size, overlap,step)

    arousal_ups = pd.Series(arousal_upsampled)  #annotation must be a pandas series, because segment_time_series expects data in this format
    aro_segments = segment_time_series(arousal_ups, window_size, overlap,step)

    time_vector_segments = segment_time_series(time_vector, window_size, overlap,step)
    
    tag_segments = segment_time_series(data_tag, window_size, overlap,step)
    


    # Verification of correct annotation overlapping
    # Verification that one chunk might contain more than one label value

    if show_plot:
        plt.figure()
        plt.plot(valence_ups, label="upsampled")
        plt.plot(val_segments[2][:], label="segment")
        plt.legend()
        plt.title("Valence")
        # plt.show(block=True)

        plt.figure()
        plt.plot(valence_ups, label="upsampled")
        plt.plot(val_segments[3][:], label="segment")
        plt.legend()
        plt.title("Valence")

        plt.figure()
        plt.plot(arousal_ups, label="upsampled")
        plt.plot(aro_segments[2][:], label="segment")
        plt.legend()
        plt.title("Arousal")
        # plt.show(block=True)

        plt.figure()
        plt.plot(arousal_ups, label="upsampled")
        plt.plot(aro_segments[3][:], label="segment")
        plt.legend()
        plt.title("Arousal")

        # check if the different pd.Series containing the arousal, valence, time, and signal chunks are syncronized. 
        chunk_number = 200
        time_vector_segments_pd = pd.Series(time_vector_segments) # pandas Series: each row contains a chunk of the time vector
        aro_segments_pd = pd.Series(aro_segments)    # pandas Series: each row contains a chunk of arousal labels

        plt.figure()
        plt.plot(time_vector_segments_pd[chunk_number], aro_segments_pd[chunk_number], label = "without time vector")
        plt.plot(aro_segments_pd[chunk_number], label = "without time vector")
        plt.legend()
        plt.title("Check if the different pd.Series containing the arousal, valence, time, and signal chunks are syncronized")


        plt.show(block=True)

    return [val_segments,aro_segments, time_vector_segments, tag_segments]



def findMajority(arr, n):
    """
    Function to find the majority element (must be an element repeated contiguously in time).
    It employes the Boyer–Moore majority vote algorithm.
    https://www.geeksforgeeks.org/boyer-moore-majority-voting-algorithm/
    """
    candidate = np.nan  # or -1
    votes = 0

    # Finding the majority candidate
    for i in range(n):
        if votes == 0:
            candidate = arr[i]
            votes = 1
        else:
            if arr[i] == candidate:
                votes += 1
            else:
                votes -= 1
    count = 0

    # Checking if the majority candidate occurs more than n/2 times
    for i in range(n):
        if arr[i] == candidate:
            count += 1
    if count > n // 2:
        return candidate
    else:
        return -1


def test_majority_voting_algorithm(annotation_segments,chunk_number):

    print("-------------------------------")
    print("test voting ALG with measured annotations:")
    plt.plot(annotation_segments[chunk_number][:], label="label segment")
    plt.legend()
    plt.show(block=True)

    n = len((annotation_segments[chunk_number][:]))
    majority = findMajority(list(annotation_segments[chunk_number][:]), n)
    print(" The majority element is :" ,majority)


def compare_annotation_chunk_decision_methods(annotation_segments, method = "all", chunk_number = 1):
    """
    Compare the different annotation chunk decision methods.
    Method = "median", "mean", "all". "all": compare mean and median.
    "Majority": It employes the Boyer–Moore majority vote algorithm. This methods is computed always, independent of the chosen method.
    """

    if method == "median":

        median = np.median(annotation_segments[chunk_number][:])
        median_array = np.ones((len(annotation_segments[chunk_number][:]),1))*median
        print("median: ", median)
        median_array_pd = pd.Series(np.ravel(median_array),index=annotation_segments[chunk_number][:].index)  # Get indexing of annotation_segments[chunk_number][:] pandas series to plot both time_series in the same plot and compare

        plt.plot(annotation_segments[chunk_number][:], label="label segment")
        plt.plot(median_array_pd, label="median")
        plt.legend()
        plt.show(block=True)

    elif method == "mean":

        mean = np.mean(annotation_segments[chunk_number][:])
        mean_array = np.ones((len(annotation_segments[chunk_number][:]),1))*mean
        print("mean: ", mean)
        mean_array_pd = pd.Series(np.ravel(mean_array),index=annotation_segments[chunk_number][:].index)  # Get indexing of annotation_segments[chunk_number][:] pandas series to plot both time_series in the same plot and compare

        plt.plot(annotation_segments[chunk_number][:], label="label segment")
        plt.plot(mean_array_pd, label="mean")
        plt.legend()
        plt.show(block=True)
    elif method == "all":

        median = np.median(annotation_segments[chunk_number][:])
        median_array = np.ones((len(annotation_segments[chunk_number][:]),1))*median
        
        median_array_pd = pd.Series(np.ravel(median_array),index=annotation_segments[chunk_number][:].index)  # Get indexing of annotation_segments[chunk_number][:] pandas series to plot both time_series in the same plot and compare
    
        mean = np.mean(annotation_segments[chunk_number][:])
        mean_array = np.ones((len(annotation_segments[chunk_number][:]),1))*mean
        
        mean_array_pd = pd.Series(np.ravel(mean_array),index=annotation_segments[chunk_number][:].index)  # Get indexing of annotation_segments[chunk_number][:] pandas series to plot both time_series in the same plot and compare

        plt.plot(annotation_segments[chunk_number][:], label="label segment")
        plt.plot(mean_array_pd, label="mean")
        plt.plot(median_array_pd, label="median")
        plt.legend()
        plt.show(block=True)    
        print("-------------------------------")
        print("median: ", median)
        print("mean: ", mean)


    print("-------------------------------")
    print("test voting ALG with measured annotations:")

    n = len((annotation_segments[chunk_number][:]))
    majority = findMajority(list(annotation_segments[chunk_number][:]), n)
    print(" The majority element is :" ,majority)



def perform_windowing(data, ppg_signals, gsr_signals, valence, arousal, seqs_order_num, participant, window_size, step, overlap, show_plot = 0):
    """
    Performes windowing / segmentation. 
    It returns a pandas DataFrame with all the segmented signals, done in a stimulus window basis. 
    Each column represents a different signal, each row represent a different window / chunk. 

    # Process:
    # 1. Baseline mean normalization (subtract A from all the time series).
    # 2. Perform segmentation on a B-windows basis (the first chunk/window start should be synced with the B-window start). See https://www.notion.so/Windowing-b8cfc88a4ab44d509d042a18ff892e0c?pvs=4#5f08d714fdb14e60bcb98709f3a4d798 
    # 3. Discard the last chunk (inside a B-window) if window_size/step doesn’t fit exactly with B-window duration.
    # 4. Discard all data corresponding to C-windows.

    if show_plot == 1, it show figures to visually check correct window overlapping and concatenation across stimuli windows

    """


    def build_dataframe_to_process(data, ppg_signals, gsr_signals, valence,arousal):
        """
        returns a pandas dataframe with filtered signals and upsampled annotations (valence and arousal)
        """
        valence_ups = nk.signal_resample(valence, sampling_rate=ANNOTATION_SAMPLING_RATE, desired_sampling_rate=SAMPLING_RATE, method="numpy")  # ups = upsampled
        valence_ups = pd.Series(valence_ups)  #annotation must be a pandas series, because segment_time_series expects data in this format
        arousal_ups = nk.signal_resample(arousal, sampling_rate=ANNOTATION_SAMPLING_RATE, desired_sampling_rate=SAMPLING_RATE, method="numpy")  # ups = upsampled
        arousal_ups = pd.Series(arousal_ups)  #annotation must be a pandas series, because segment_time_series expects data in this format

        data_cleaned = data.copy()  # ppg and gsr filtered copy
        data_cleaned['bvp'] = ppg_signals['PPG_Clean']
        data_cleaned['gsr'] = gsr_signals['EDA_Clean']
        data_cleaned['val'] = valence_ups
        data_cleaned['aro'] = arousal_ups

        return data_cleaned



    def obtain_stimuli_sequence_order (seqs_order_num):
        """
        Returns the stimuli sequence order of each participant  (dropping the neutral screens)
        """
        # Identify rows containing neutral screens (start video, end video, blue screen)
        mask_VIDEO_ID_10 = (seqs_order_num == VIDEO_ID_10).any(axis=1)
        mask_VIDEO_ID_11 = (seqs_order_num == VIDEO_ID_11).any(axis=1)
        mask_VIDEO_ID_12 = (seqs_order_num == VIDEO_ID_12).any(axis=1)

        mask = mask_VIDEO_ID_10 | mask_VIDEO_ID_11 | mask_VIDEO_ID_12 # contains all neutral screens

        filtered_seqs_order_num = seqs_order_num[~mask]  #contains stimuli sequence order (without neutral screens)

        return filtered_seqs_order_num


    def combine_chunks(df_base, segments_pd, video_id, time_series_name):
        """
        Combine chunks in order, taking into account all the stimulus windows in order.
    
        segments_pd = pandas Series
        video_id = constant number
        if add_video = 1, it adds a column to the pd.Series with the video_id number. 
        time_series_name is a string with the time series name. 
        """

        # Create a new Pandas Series with the video-ID (constant number)
        video_column = pd.Series( [video_id] * len(segments_pd))

        # Concatenate the original Series and the new Series along the columns
        combined_df = pd.concat([segments_pd, video_column], axis=COLUMNS)

        # Rename the columns:
        combined_df.columns = [time_series_name + '_' +'chunks', 'video']
        # print("header combined_df: ", list(combined_df.columns))

        # build a dataFrame with all the chunks of all the stimulus windows together
        if df_base.equals(segments_pd):
            df = combined_df.copy()
        else:
            df = pd.concat([df_base, combined_df], axis = INDEX)
    
        # Display the resulting DataFrame
        # print(df)

        return df

    def verify_correct_overlapping(data_cleaned, gsr_segments, ppg_segments, val_segments, aro_segments, time_vector_segments, tag_segments):
        # Verification of correct overlapping
        # last chunk of data doesn't fit in a window (because window_size and step doesn't match with stimulus window duration)

        # check GSR overlapping:
        plt.plot(data_cleaned['gsr'], label="Cleaned GSR")
        plt.plot((data_cleaned['tag']-data_cleaned['tag'].mean())*data_cleaned['gsr'].max()+data_cleaned['gsr'].mean(), label="stimulus = high | neutral = low") #plots tag over signal: useful to check correct windowing
        plt.plot(stimulus_window['gsr'])
        for i in range(np.shape(gsr_segments)[0]):
            plt.plot(gsr_segments[i])
        plt.title('GSR')
        plt.legend(loc='upper right')
        plt.show(block=True)


        # check PPG overlapping:
        # last chunk of data doesn't fit in a window (because window_size and step doesn't match with stimulus window duration)
        plt.plot(data_cleaned['bvp'], label="Cleaned PPG")
        plt.plot((data_cleaned['tag']-data_cleaned['tag'].min())*data_cleaned['bvp'].max(), label="stimulus = high | neutral = low") #plots tag over signal: useful to check correct windowing
        plt.plot(stimulus_window['bvp'])
        for i in range(np.shape(ppg_segments)[0]):
            plt.plot(ppg_segments[i])
        plt.title('PPG')
        plt.legend(loc='upper right')
        plt.show(block=True)

        # check Annotations overlapping:
        # last chunk of data doesn't fit in a window (because window_size and step doesn't match with stimulus window duration)
        plt.plot(data_cleaned['val'], label="Valence")
        plt.plot((data_cleaned['tag']-data_cleaned['tag'].mean())*data_cleaned['val'].max()+data_cleaned['val'].mean(), label="stimulus = high | neutral = low") #plots tag over signal: useful to check correct windowing
        for i in range(np.shape(val_segments)[0]):
            plt.plot(val_segments[i])
        plt.title('Valence')
        plt.legend(loc='upper right')
        plt.show(block=True)

        # check time-vector overlapping:
        # last chunk of data doesn't fit in a window (because window_size and step doesn't match with stimulus window duration)
        plt.plot(data_cleaned['daqtime'], label="Time-vector")
        plt.plot((data_cleaned['tag']-data_cleaned['tag'].mean())*data_cleaned['daqtime'].max()+data_cleaned['daqtime'].mean(), label="stimulus = high | neutral = low") #plots tag over signal: useful to check correct windowing
        for i in range(np.shape(time_vector_segments)[0]):
            plt.plot(time_vector_segments[i])
        plt.title('Time-vector')
        plt.legend(loc='upper right')
        plt.show(block=True)

        # check tag overlapping:
        # last chunk of data doesn't fit in a window (because window_size and step doesn't match with stimulus window duration)
        plt.plot(data_cleaned['tag'], label="Tag")
        plt.plot((data_cleaned['tag']-data_cleaned['tag'].mean())*data_cleaned['tag'].max()+data_cleaned['tag'].mean(), label="stimulus = high | neutral = low") #plots tag over signal: useful to check correct windowing
        for i in range(np.shape(tag_segments)[0]):
            plt.plot(tag_segments[i])
        plt.title('Tag')
        plt.legend(loc='upper right')
        plt.show(block=True)

    def build_unique_dataframe(gsr_df, ppg_df, val_df, aro_df, time_df, tag_df):
        """
        Builds a unique DataFrame with all time-series: data with all chunks of all the stimulus windows together 
        (each biosignal, annotation will be a separate column)
        Returns a dataFrame with this header: 
        ['gsr_chunks', 'ppg_chunks', 'val_chunks', 'aro_chunks', 'time_chunks', 'tag_chunks', 'video']
        """

        # build a unique DataFrame with all time-series: data with all chunks of all the stimulus windows together (each biosignal, annotation will be a separate column)
        combined_df = pd.concat([gsr_df, ppg_df, val_df, aro_df, time_df, tag_df], axis=COLUMNS)

        # Remove duplicated 'video' columns except one
        current_columns = list(combined_df.columns)
        current_columns[1] = 'video_id'
        combined_df.columns = current_columns
        list(combined_df.columns)

        # Find duplicate column names
        duplicate_columns = combined_df.columns[combined_df.columns.duplicated()]

        # Drop columns with duplicate names
        combined_df = combined_df.drop(columns=duplicate_columns)
        list(combined_df.columns)

        # Move the 'video' column to the end
        video_column = combined_df.pop('video_id')
        combined_df['video'] = video_column
        print(list(combined_df.columns))
        
        return combined_df
    
    def verify_correct_concatenation(data_cleaned, gsr_df, ppg_df):
        # check correct concatenation:
        # check GSR overlapping:
        plt.plot(data_cleaned['gsr'], label="Cleaned GSR")
        plt.plot((data_cleaned['tag']-data_cleaned['tag'].mean())*data_cleaned['gsr'].max()*2+data_cleaned['gsr'].mean(), label="stimulus = high | neutral = low") #plots tag over signal: useful to check correct windowing
        # plt.plot(stimulus_window['gsr'])
        # plt.plot(gsr_df['chunks'].iloc[100])
        for i in range(gsr_df['gsr_chunks'].shape[0]):
            plt.plot(gsr_df['gsr_chunks'].iloc[i])
        plt.title('Check correct concatenation - GSR')
        plt.legend(loc='upper right')
        plt.show(block=True)


        # check correct concatenation:
        # check PPG overlapping:
        plt.plot(data_cleaned['bvp'], label="Cleaned PPG")
        plt.plot((data_cleaned['tag']-data_cleaned['tag'].min())*data_cleaned['bvp'].max()*1, label="stimulus = high | neutral = low") #plots tag over signal: useful to check correct windowing
        for i in range(ppg_df['ppg_chunks'].shape[0]):
            plt.plot(ppg_df['ppg_chunks'].iloc[i])
        plt.title('Check correct concatenation - PPG')
        plt.legend(loc='upper right')
        plt.show(block=True)


    data_cleaned = build_dataframe_to_process(data, ppg_signals, gsr_signals, valence,arousal)
    filtered_seqs_order_num = obtain_stimuli_sequence_order (seqs_order_num)
    
    # Loop over all 8 stimuli (each stimulus is then windowed)
    for i in range(filtered_seqs_order_num.shape[0]):
    # for i in range(1):
        # one particular stimulus window with all its data (physiological, val, aro, etc):
        stimulus_window = data_cleaned[data_cleaned['video'] == filtered_seqs_order_num['sub_' + participant].iloc[i]]    # contains one particular B-Window

        gsr_segments = segment_time_series(stimulus_window['gsr'], window_size, overlap,step)  # gsr_segments for one particular stimulus window
        ppg_segments = segment_time_series(stimulus_window['bvp'], window_size, overlap,step)
        val_segments = segment_time_series(stimulus_window['val'], window_size, overlap,step)
        aro_segments = segment_time_series(stimulus_window['aro'], window_size, overlap,step)
        time_vector_segments = segment_time_series(stimulus_window['daqtime'], window_size, overlap,step)
        tag_segments = segment_time_series(stimulus_window['tag'], window_size, overlap,step)

        # if show_plot == 1:
        #     verify_correct_overlapping(data_cleaned, gsr_segments, ppg_segments, val_segments, aro_segments, time_vector_segments, tag_segments)
        
        gsr_segments_pd = pd.Series(gsr_segments)
        ppg_segments_pd = pd.Series(ppg_segments)
        val_segments_pd = pd.Series(val_segments)
        aro_segments_pd = pd.Series(aro_segments)
        time_vector_segments_pd = pd.Series(time_vector_segments)
        tag_segments_pd = pd.Series(tag_segments)

        # build a dataFrame with all chunks of all the stimulus windows together (one dataFrame for each time series)
        if i == 0:
            gsr_df = combine_chunks(gsr_segments_pd, gsr_segments_pd, filtered_seqs_order_num['sub_' + participant].iloc[i], 'gsr')
            ppg_df = combine_chunks(ppg_segments_pd, ppg_segments_pd, filtered_seqs_order_num['sub_' + participant].iloc[i], 'ppg')
            val_df = combine_chunks(val_segments_pd, val_segments_pd, filtered_seqs_order_num['sub_' + participant].iloc[i], 'val')
            aro_df = combine_chunks(aro_segments_pd, aro_segments_pd, filtered_seqs_order_num['sub_' + participant].iloc[i], 'aro')
            time_df = combine_chunks(time_vector_segments_pd, time_vector_segments_pd, filtered_seqs_order_num['sub_' + participant].iloc[i], 'time')
            tag_df = combine_chunks(tag_segments_pd, tag_segments_pd, filtered_seqs_order_num['sub_' + participant].iloc[i], 'tag')
        else:
            gsr_df = combine_chunks(gsr_df, gsr_segments_pd, filtered_seqs_order_num['sub_' + participant].iloc[i], 'gsr')
            ppg_df = combine_chunks(ppg_df, ppg_segments_pd, filtered_seqs_order_num['sub_' + participant].iloc[i], 'ppg')
            val_df = combine_chunks(val_df, val_segments_pd, filtered_seqs_order_num['sub_' + participant].iloc[i], 'val')
            aro_df = combine_chunks(aro_df, aro_segments_pd, filtered_seqs_order_num['sub_' + participant].iloc[i], 'aro')
            time_df = combine_chunks(time_df, time_vector_segments_pd, filtered_seqs_order_num['sub_' + participant].iloc[i], 'time')
            tag_df = combine_chunks(tag_df, tag_segments_pd, filtered_seqs_order_num['sub_' + participant].iloc[i], 'tag')

    if show_plot == 1:
        verify_correct_concatenation(data_cleaned, gsr_df, ppg_df)
        
    combined_df = build_unique_dataframe(gsr_df, ppg_df, val_df, aro_df, time_df, tag_df)

    return combined_df

def verify_performed_windowing(combined_df, gsr_signals):
    """
    It verifies the windowing operation by showing a particular signal
    and its chunks
    """
    plt.plot(gsr_signals['EDA_Clean'], label="GSR chunk")
    gsr_min_value = min(gsr_signals['EDA_Clean'])
    gsr_max_value = max(gsr_signals['EDA_Clean'])
    gsr_scaled_tag = [scale_tag(i,gsr_min_value,gsr_max_value) for i in data['tag']] 
    gsr_scaled_tag = pd.Series(gsr_scaled_tag)
    plt.plot((gsr_scaled_tag-gsr_scaled_tag.mean())*gsr_signals['EDA_Clean'].max()+gsr_signals['EDA_Clean'].mean(), label="stimulus = high | neutral = low") #plots tag over signal: useful to check correct windowing
        
    for i in range(combined_df['gsr_chunks'].shape[0]):
        plt.plot(combined_df['gsr_chunks'].iloc[i])
    plt.title('Check correct windowing - GSR')
    plt.legend(loc='upper right')
    plt.show(block=True)



def map_labels_bipartite(x, number_thresholds = 1, threshold = 5, L = 3, H = 7):  
    """
    It performs a binary mapping of labels. 
    Maps valence and arousal (from 1.0 to 9.0 (float) scale to a binary scale: 0 or 1)
    It runs in one of two modes:

    1) Classic: one threshold is employ to do the mapping.
    (typically, threshold = 5) 
    if number_thresholds = 1, then
    map x -> 0  |    x <= threshold    
    map x -> 1  |    x > threshold

    2) Discard: two thresholds are employed to do the mapping: L (low threshold)
    and H (high threshold)
        
    If number_thresholds = 2, 
    map x -> 0  |    x <= L    
    map x -> 1  |    x >= H
    discard x |   L < x < H    
    """

    if number_thresholds == 1:

        if x <= threshold:
            return 0
        elif x > threshold:
            return 1
        
    elif number_thresholds == 2:
        if x <= L:
            return 0
        elif x >= H:
            return 1
        else: 
            return np.nan



def map_labels_tripartite(x, L = 3, H = 7):  
    """
    It performs a binary mapping of labels. 
    Maps valence and arousal (from 1.0 to 9.0 (float) scale to a three class: 0, 1, 11)
        
    Two thresholds are employed to do the mapping: L (low threshold)
    and H (high threshold). 
        
    map x -> 0  |    x <= L    
    map x -> 1  |    L < x < H
    map x -> 2  |    x >= H
            
    """
    if x <= L:
        return LOW_LABEL_MAP_VALUE
    elif L < x < H:
        return MEDIUM_LABEL_MAP_VALUE    
    elif x >= H:
        return HIGH_LABEL_MAP_VALUE


def map_valence_arousal(combined_df, scheme, number_thresholds, threshold, L, H):
    """
    Map valence and arousal from a pandas datrafeme, from float 1.0 to 9.0 to a two or three class label
    depending on the chosen scheme = 'bipartite' or 'tripartite'.

    """
    combined_df_copy = combined_df.copy()

    labels = ['val_chunks','aro_chunks']

    if scheme == 'bipartite':

        for label in range(len(labels)): 
            for index in range(combined_df_copy[labels[label]].shape[0]):
                combined_df_copy[labels[label]].iloc[index] = combined_df_copy[labels[label]].iloc[index].apply(map_labels_bipartite, args=(number_thresholds, threshold, L, H))
            

    elif scheme == 'tripartite':

        for label in range(len(labels)): 
            for index in range(combined_df_copy[labels[label]].shape[0]):
                combined_df_copy[labels[label]].iloc[index] = combined_df_copy[labels[label]].iloc[index].apply(map_labels_tripartite, args=(L, H))
           

    return combined_df_copy



def check_correct_mapping(combined_df, combined_df_median, combined_df_mapped):
    """
    Check if mapping to a whole column of the dataframe worked
    """
    
    index = 10

    labels = ['val_chunks','aro_chunks']

    label = labels[0]


    print("verify_ check_correct_mapping (INSIDE running method): -------------")
    print("raw: -------------")
    print(combined_df[label].iloc[index])
    print("Median: -------------")
    print(combined_df_median[label].iloc[index])
    print("Mapped: -------------")
    print(combined_df_mapped[label].iloc[index])

    plt.plot(combined_df[label].iloc[index], label='raw')
    plt.plot(combined_df_median[label].iloc[index], label='median')
    plt.plot(combined_df_mapped[label].iloc[index], label='mapped')
    plt.legend()
    plt.title("Verify / check correct mapping")
    plt.show(block=True)


def median_voting(combined_df):
    """
    Replaces labels with their median on a chunk basis.
    Each chunk will have a constant number: its median.
    This voting should be applied before mapping the labels (employing bipartition or tripartition scheme).
    """

    def median_voting_mapping(x, median):
        return median

    combined_df_copy = combined_df.copy()

    labels = ['val_chunks','aro_chunks']

    for label_index in range(len(labels)):
        for i in range(combined_df_copy[labels[label_index]].shape[0]):
            median = combined_df_copy[labels[label_index]].iloc[i].median()
            combined_df_copy[labels[label_index]].iloc[i] = combined_df_copy[labels[label_index]].iloc[i].apply(median_voting_mapping, args=(median,))
            

    return combined_df_copy


def verify_median_voting(combined_df, combined_df_median):
    """
    Verify that median_voting method works fine.
    """

    index = 0
    labels = ['val_chunks','aro_chunks']
    label = labels[1]

    print("verify_median_voting: -------------")
    print("raw: -------------")
    print(combined_df[label].iloc[index])
    print("Median: -------------")
    print(combined_df_median[label].iloc[index])
    plt.plot(combined_df[label].iloc[index], label='raw')
    plt.plot(combined_df_median[label].iloc[index], label='median voting')
    plt.legend()
    plt.title("verify_median_voting")
    plt.show(block=True)



def assign_unique_label(y_windowed):
    """
    Build a new dataframe with a unique binary label for each window 
    and assigns True / False to each 0 / 1 value. 
    """
    # Build a new dataframe with a unique label for each window 

    # Initialize an empty DataFrame with the same shape as y_windowed
    y_median = pd.DataFrame(index=range(y_windowed.shape[0]), columns=y_windowed.columns)

    # Iterate over each row and column of y_windowed
    for i in range(y_windowed.shape[0]):
        for column in y_windowed.columns:
            # Extract the first element of the time series in the current element
            first_element = y_windowed.loc[i, column].iloc[0]
            # Assign the first element to the corresponding location in the new DataFrame
            y_median.loc[i, column] = first_element

    # Verify the shape of the new DataFrame
    # print(y_median.shape)  # Output should be (418, 2)

    # convert labels = [0,1] (dtype=object) to boolean (True, False)
    y_median = (y_median != 0).astype(bool)

    return y_median



def build_dataset(combined_df):
    """
    Get dataset into X and y pandas DataFrames
    
    combined_df: a pandas dataframe with windowed data and labels. 
    Keys:
    Index(['gsr_chunks', 'ppg_chunks', 'val_chunks', 'aro_chunks', 'time_chunks',
       'video'], dtype='object') 

    returns: X_windowed, y_median
    
    X_windowed keys: Index(['gsr_chunks', 'ppg_chunks'], dtype='object' 
    shape: (number of windows, number of modalities / physiological signals)
    
    y_median keys: Index(['val_chunks', 'aro_chunks'], dtype='object')
    shape: (number of windows, number of different types of labels)
    """

    # The data matrix X
    X = combined_df.loc[:,'gsr_chunks':'ppg_chunks']
    # Labels
    y = combined_df.loc[:,'val_chunks':'aro_chunks']

    # Reset the indices of X and y to have clear access to each element.
    X_windowed = X.copy()
    X_windowed.reset_index(drop=True, inplace =  True)
    # Xc is a time-series, not a feature vector / matrix

    y_windowed = y.copy()
    y_windowed.reset_index(drop=True, inplace =  True)

    # Build a new dataframe with a unique binary label for each window and assigns True / False to each 0 / 1 value. 
    y_median = assign_unique_label(y_windowed)  

    return X_windowed, y_median



def poincare_indices(data, dt=1, show=False):
    """
    Compute poincaré indices
    - **SD1**: Standard deviation perpendicular to the line of identity. It is an index of short-term RR interval fluctuations, i.e., beat-to-beat variability. It is equivalent (although on another scale) to RMSSD, and therefore it is redundant to report correlation with both.
    - **SD2**: Standard deviation along the identity line. Index of long-term HRV changes.
    - **SD1/SD2**: ratio of *SD1* to *SD2*. Describes the ratio of short term to long term variations in HRV.
    - **S**: Area of ellipse described by *SD1* and *SD2* (`pi * SD1 * SD2`). It is proportional to *SD1SD2*.
    
    data: pandas series
    dt = delta 't'. When drawing a Poincaré plot, 'x(n+dt)' is plotted against 'x(n)', where dt represents the shift.
    
    If show = True, it draws the poincaré plot.
    """ 
    out={}

    # Create a Poincaré plot and compute its indices
    if show:
        plt.scatter(data[:-dt], data[dt:])
        plt.title('Poincaré Plot')
        plt.xlabel('Previous Value')
        plt.ylabel('Current Value')
        plt.show()

    # Poincaré indices:
    # Golińska, A.K.: Poincaré plots in analysis of selected biomedical signals. Stud. Logic, Gramm. Rhetor. 35, 48, 117–127 (2013). https://doi.org/10.2478/slgr-2013-0031.
    # Calculate SD1 and SD2 (standard deviations)
    sd1 = np.std(data[:-dt] - data[dt:]) / np.sqrt(2)
    sd2 = np.std(data[:-dt] + data[dt:]) / np.sqrt(2) # given by chatGPT
    # # it shows the formulas given by Golińska and chapGPT, are equivalent
    # sd2v2 = np.sqrt(2*np.std(data[:-dt])**2-0.5*np.std(data[:-dt] - data[dt:])**2)  #given by Golińska
    
    area = np.pi*sd1*sd2

    # Calculate and print the SD1/SD2 ratio
    sd1_sd2_ratio = sd1 / sd2
    
    out["SD1"] = sd1
    out["SD2"] = sd2

    # SD1 / SD2
    out["SD12"] = sd1 / sd2

    # Area of ellipse described by SD1 and SD2
    out["S"] = area

    # out["SDXX"] = np.sqrt((sd1**2+sd2**2)/5)  # Resolution: I won't extract this feature because (in general) it has low variability 

    return out

def check_poincare_plots(X_windowed, show = False):
    """ 
    Poincaré plot for PPG signal (repeating biosignal):
    It show that Poincaré indices varies over the different lags
    """
    
    time_series = X_windowed.loc[:,'ppg_chunks'].iloc[10]
    for lag in range(1,10,1):
        plt.figure()
        print(poincare_indices(np.array(time_series),dt=lag,show=show)) # poincare_indices method was previously tested above. 
        # plt.figure()
        # lle, info = nk.complexity_lyapunov(time_series, delay=lag, method="rosenstein", show=show)
        # print("LLE: ", lle)
        # print(info)


def feature_extract_GSR_PPG_non_linear(X_windowed):
    """
    Extract features (in a window-based regime)
    X_windowed: a pandas DatraFrame contaning the time-series windows as rows and gsr and ppg data as columns. 
    return: feature vector (pandas DataFrame) with this features:
    [pi_gsr_l1['SD1'], pi_gsr_l1['SD2'], pi_gsr_l1['SD12'], pi_gsr_l1['S'],pi_gsr_l10['SD1'], pi_gsr_l10['SD2'], pi_gsr_l10['SD12'], pi_gsr_l10['S'], le_gsr, apen_gsr, pi_ppg_l1['SD1'], pi_ppg_l1['SD2'], pi_ppg_l1['SD12'], pi_ppg_l1['S'],pi_ppg_l10['SD1'], pi_ppg_l10['SD2'], pi_ppg_l10['SD12'], pi_ppg_l10['S'],le_ppg, apen_ppg ]
    shape: (number of windows, number of features)

    pi: poincaré indices
    - **SD1**: Standard deviation perpendicular to the line of identity. It is an index of short-term RR interval fluctuations, i.e., beat-to-beat variability. It is equivalent (although on another scale) to RMSSD, and therefore it is redundant to report correlation with both.
    - **SD2**: Standard deviation along the identity line. Index of long-term HRV changes.
    - **SD1/SD2**: ratio of *SD1* to *SD2*. Describes the ratio of short term to long term variations in HRV.
    - **S**: Area of ellipse described by *SD1* and *SD2* (`pi * SD1 * SD2`). It is proportional to *SD1SD2*.
    
    gsr / ppg: GSR / PPG signal
    li: lag i
    le: Lyapunov Exponent
    apen: ApEn: Approximate Entropy

    """
    
    print("----- Extracting features...")    
    
    fv = [] #feature vector 

    nbr_instances = X_windowed.shape[0]

    for i in range(nbr_instances):

        ppg_time_series = X_windowed.loc[:,'ppg_chunks'].iloc[i].values
        gsr_time_series = X_windowed.loc[:,'gsr_chunks'].iloc[i].values # return numpy array (poincare indices expects this format)

        lag = 1
        pi_gsr_l1 = poincare_indices(gsr_time_series,dt=lag,show=False)
        pi_ppg_l1 = poincare_indices(ppg_time_series,dt=lag,show=False)

        lag = 10
        pi_gsr_l10 = poincare_indices(gsr_time_series,dt=lag,show=False)
        pi_ppg_l10 = poincare_indices(ppg_time_series,dt=lag,show=False)

        le_gsr, parameters = nk.complexity_lyapunov(gsr_time_series, method="rosenstein", show=False)
        le_ppg, parameters = nk.complexity_lyapunov(ppg_time_series, method="rosenstein", show=False)

        apen_gsr, parameters = nk.entropy_approximate(gsr_time_series, corrected=False)
        apen_ppg, parameters = nk.entropy_approximate(ppg_time_series, corrected=False)

        fv.append([pi_gsr_l1['SD1'], pi_gsr_l1['SD2'], pi_gsr_l1['SD12'], pi_gsr_l1['S'],pi_gsr_l10['SD1'], pi_gsr_l10['SD2'], pi_gsr_l10['SD12'], pi_gsr_l10['S'], le_gsr, apen_gsr, pi_ppg_l1['SD1'], pi_ppg_l1['SD2'], pi_ppg_l1['SD12'], pi_ppg_l1['S'],pi_ppg_l10['SD1'], pi_ppg_l10['SD2'], pi_ppg_l10['SD12'], pi_ppg_l10['S'],le_ppg, apen_ppg ])

    fv = np.array(fv)  # fv = feature vector

    # Build feature vector as a DataFrame
    d = {'g_l1_SD1': fv[:,0],'g_l1_SD2': fv[:,1], 'g_l1_SD12': fv[:,2], 'g_l1_S': fv[:,3], 'g_l10_SD1': fv[:,4],'g_l10_SD2': fv[:,5], 'g_l10_SD12': fv[:,6], 'g_l10_S': fv[:,7], 'le_gsr': fv[:,8],'apen_gsr': fv[:,9], 'p_l1_SD1': fv[:,10],'p_l1_SD2': fv[:,11], 'p_l1_SD12': fv[:,12], 'p_l1_S': fv[:,13], 'p_l10_SD1': fv[:,14],'p_l10_SD2': fv[:,15], 'p_l10_SD12': fv[:,16], 'p_l10_S': fv[:,17], 'le_ppg': fv[:,18],'apen_ppg': fv[:,19]}
    # d = {'SD1': fv[:,0],'SD2': [poinc_ind['SD2']],'SD12': [poinc_ind['SD12']],'S': [poinc_ind['S']]}
    xf = pd.DataFrame(data=d)  # feature vector (in DataFrame format)
    
    return xf


    
def evaluate_alg_CASE(model,xf,y_median, report=False):
    """
    Evaluate model (arousal and valence separatedly)
    model = KNeighborsClassifier, DecisionTreeClassifier, etc 
    xf = complete dataset (pandas dataframe)
    y_median = valence and arousal median (pandas dataframe)
    report = if True, print results.

    returns: np.array with performance metrics for this subject:
    
    np.array([(val_cohen, val_uar, val_acc,val_gm, val_f1, aro_cohen, aro_uar,aro_acc,aro_gm, aro_f1)])
    """
    
    print("------------- Evaluating model --------------")

    # define the pipeline 
    steps = list()
    # steps.append(('scaler', MinMaxScaler(feature_range=(-1,1))))
    steps.append(('scaler2', StandardScaler()))
    steps.append(('model', model))
    
    pipeline = Pipeline(steps=steps)
    # define the evaluation procedure
    # cv = RepeatedStratifiedKFold(n_splits=10, n_repeats=5, random_state=1)
    cv = RepeatedStratifiedKFold(n_splits=10, n_repeats=1, random_state=1)
    # evaluate the model using cross-validation
    kappa_scorer = make_scorer(cohen_kappa_score) # Define the Cohen's kappa scorer

    gmean_scorer = make_scorer(geometric_mean_score) # Define the Cohen's kappa scorer

    scoring = ['accuracy', 'balanced_accuracy', 'f1_macro','roc_auc']  # names from https://scikit-learn.org/stable/modules/model_evaluation.html#classification-metrics 
    
    # VALENCE model evaluation:
    scores = cross_validate(pipeline, xf, y_median.loc[:,'val_chunks'].values, scoring=scoring, cv=cv, n_jobs=-1)
    scores_kappa = cross_val_score(pipeline, xf, y_median.loc[:,'val_chunks'].values, scoring=kappa_scorer, cv=cv, n_jobs=-1)
    scores_gmean = cross_val_score(pipeline, xf, y_median.loc[:,'val_chunks'].values, scoring=gmean_scorer, cv=cv, n_jobs=-1)

    val_cohen = mean(scores_kappa)
    val_uar = mean(scores['test_balanced_accuracy'])
    val_acc = mean(scores['test_accuracy'])
    val_gm = mean(scores_gmean)
    val_f1 = mean(scores['test_f1_macro'])
    
    # Arousal model evaluation:
    scores = cross_validate(pipeline, xf, y_median.loc[:,'aro_chunks'].values, scoring=scoring, cv=cv, n_jobs=-1)
    scores_kappa = cross_val_score(pipeline, xf, y_median.loc[:,'aro_chunks'].values, scoring=kappa_scorer, cv=cv, n_jobs=-1)
    scores_gmean = cross_val_score(pipeline, xf, y_median.loc[:,'aro_chunks'].values, scoring=gmean_scorer, cv=cv, n_jobs=-1)

    aro_cohen = mean(scores_kappa)
    aro_uar = mean(scores['test_balanced_accuracy'])
    aro_acc = mean(scores['test_accuracy'])
    aro_gm = mean(scores_gmean)
    aro_f1 = mean(scores['test_f1_macro'])

    if report:
        # report Valence performance
        print("Valence: -----")
        print('kappa: %.3f (%.3f)' % (val_cohen*100, std(scores_kappa)*100))
        print('balanced_accuracy (UAR): %.3f (%.3f)' % (val_uar*100, std(scores['test_balanced_accuracy'])*100))
        print('Accuracy: %.3f (%.3f)' % (val_acc*100, std(scores['test_accuracy'])*100))
        print('g-mean: %.3f (%.3f)' % (val_gm*100, std(scores_gmean)*100))
        print('f1_macro: %.3f (%.3f)' % (val_f1*100, std(scores['test_f1_macro'])*100))
        print('roc_auc: %.3f (%.3f)' % (mean(scores['test_roc_auc'])*100, std(scores['test_roc_auc'])*100))
        

        # report Arousal performance
        print("Arousal: -----")
        print('kappa: %.3f (%.3f)' % (aro_cohen*100, std(scores_kappa)*100))
        print('balanced_accuracy (UAR): %.3f (%.3f)' % (aro_uar*100, std(scores['test_balanced_accuracy'])*100))
        print('Accuracy: %.3f (%.3f)' % (aro_acc*100, std(scores['test_accuracy'])*100))
        print('g-mean: %.3f (%.3f)' % (aro_gm*100, std(scores_gmean)*100))
        print('f1_macro: %.3f (%.3f)' % (aro_f1*100, std(scores['test_f1_macro'])*100))
        print('roc_auc: %.3f (%.3f)' % (mean(scores['test_roc_auc'])*100, std(scores['test_roc_auc'])*100))
        
    return np.array([(val_cohen, val_uar, val_acc,val_gm, val_f1, aro_cohen, aro_uar,aro_acc,aro_gm, aro_f1)])


def format_dataset_for_DL(X_windowed, y_median):
    """
    Convert the dataset to the deep learning algorithm's expected shape 
    and format, both for the time series and labels.
    """

    # Physiological signals
    gsr_ts_arousal = X_windowed.loc[:,"gsr_chunks"].tolist() # conver time-series to a list of arrays
    gsr_ts_valence = X_windowed.loc[:,"gsr_chunks"].tolist()
    ppg_ts_arousal = X_windowed.loc[:,"ppg_chunks"].tolist()
    ppg_ts_valence = X_windowed.loc[:,"ppg_chunks"].tolist()

    # Convert the list of arrays into a numpy array
    gsr_ts_arousal = np.array(gsr_ts_arousal)
    gsr_ts_valence = np.array(gsr_ts_valence)
    ppg_ts_arousal = np.array(ppg_ts_arousal)
    ppg_ts_valence = np.array(ppg_ts_valence)

    # Deep Learning architecture needs feature vectors with shapes, e.g, (74,20,1), where the second dimension has the time series data
    gsr_ts_arousal=np.reshape(gsr_ts_arousal,(np.shape(gsr_ts_arousal)[0],np.shape(gsr_ts_arousal)[1],1))
    ppg_ts_arousal=np.reshape(ppg_ts_arousal,(np.shape(ppg_ts_arousal)[0],np.shape(ppg_ts_arousal)[1],1))
    gsr_ts_valence=np.reshape(gsr_ts_valence,(np.shape(gsr_ts_valence)[0],np.shape(gsr_ts_valence)[1],1))
    ppg_ts_valence=np.reshape(ppg_ts_valence,(np.shape(ppg_ts_valence)[0],np.shape(ppg_ts_valence)[1],1))

    # labels should be numpy arrays
    valence = y_median.loc[:,'val_chunks'].values
    arousal = y_median.loc[:,'aro_chunks'].values

    return gsr_ts_arousal, ppg_ts_arousal, gsr_ts_valence, ppg_ts_valence, valence, arousal


def imbalance_test(valence, arousal, show_plots = False):

    print("Class Imbalance test (arousal)....")
    arousal_d_one_class = class_imbalance_test(arousal,show_plots)
    gsr_arousal_d_one_class = arousal_d_one_class

    print("Class Imbalance test (valence)....")
    valence_d_one_class = class_imbalance_test(valence,show_plots)
    gsr_valence_d_one_class = valence_d_one_class

    return arousal_d_one_class, gsr_arousal_d_one_class, valence_d_one_class, gsr_valence_d_one_class


def alg_performance_eval(xf, y_median, knn, dt, rf, svm, gbm, arousal_d_one_class, valence_d_one_class, gsr_arousal_d_one_class, gsr_valence_d_one_class, gsr_ts_arousal, ppg_ts_arousal, arousal, gsr_ts_valence, ppg_ts_valence, valence):
    """
    Test different algorithms performance on dataset.
    xf : (windowed) dataset (pandas DataFrame). Run xf.keys() to find the features names.
    y_median: labels (pandas DataFrame). Run y_median.keys() to find the features names.

    returns:  performance (numpy array) with metrics.
    Metrics order:
    [(val_cohen, val_uar, val_acc,val_gm, val_f1, aro_cohen, aro_uar,aro_acc,aro_gm, aro_f1)]
    ALG order: 
    knn = 0; dt = 1; rf = 2; svm = 3; gbm = 4; BDDAE = 5; dummy = 5
    """

    performance = np.array([(0.0, 0.0,0.0, 0.0, 0.0, 0.0,0.0,0.0,0.0,0.0)]) # store ALG metrics
    # perfomance ALG metrics order:
    # [(val_cohen, val_uar, val_acc,val_gm, val_f1, aro_cohen, aro_uar,aro_acc,aro_gm, aro_f1)]
    # AGL : 
    # knn = 0; dt = 1; rf = 2; svm = 3; gbm = 4; BDDAE = 5; dummy = 5"

    # Shallow Learning performance evaluation
    if arousal_d_one_class and valence_d_one_class:
        performance = np.empty((5,10))  # 5: number of SL ALG (without Dummy Clf); 10: number of metrics
        performance [:] = np.nan
    else:
        # ------- KNN ------------
        performance[-1] = evaluate_alg_CASE(knn,xf,y_median, report=False)
        # ------- DT ------------
        performance = np.vstack([performance,evaluate_alg_CASE(dt,xf,y_median, report=False)])
        # ------- RF -----------------
        performance = np.vstack([performance,evaluate_alg_CASE(rf,xf,y_median, report=False)])
        # -------- SVM -------------------
        performance = np.vstack([performance,evaluate_alg_CASE(svm,xf,y_median, report=False)])
        #  ------ GBM ---------------
        performance = np.vstack([performance,evaluate_alg_CASE(gbm,xf,y_median, report=False)])
    
    # Deep Learning performance evaluation

    if not(gsr_arousal_d_one_class) and not(gsr_valence_d_one_class):
        #  ------ BDDAE ---------------
        # del model
        performance = np.vstack([performance,evaluate_alg_DL(gsr_ts_arousal, ppg_ts_arousal, arousal, gsr_ts_valence, ppg_ts_valence, valence)])

        # [x_trainGSR,x_testGSR, x_trainPPG, x_testPPG, y_train, y_test] = split_ds(gsr_ts_arousal, ppg_ts_arousal, gsr_arousal_d)

    elif not(gsr_arousal_d_one_class) and gsr_valence_d_one_class:
        #  ------ BDDAE ---------------
        # del model
        performance = np.vstack([performance,evaluate_alg_DL_arousal(gsr_ts_arousal, ppg_ts_arousal, arousal, gsr_ts_valence, ppg_ts_valence, valence)])

    elif gsr_arousal_d_one_class and not(gsr_valence_d_one_class):
        #  ------ BDDAE ---------------
        # del model
        performance = np.vstack([performance,evaluate_alg_DL_valence(gsr_ts_arousal, ppg_ts_arousal, arousal, gsr_ts_valence, ppg_ts_valence, valence)])

    elif gsr_arousal_d_one_class and gsr_valence_d_one_class:
        performance_this_case = np.empty((1,10))
        performance_this_case [:] = np.nan
        performance = np.vstack([performance, performance_this_case])

    
    #  ------ Dummy Classifier -------------------

    model =  DummyClassifier(strategy="most_frequent")

    # Shallow Learning performance evaluation
    if arousal_d_one_class and valence_d_one_class:
        performance_this_case = np.empty((1,10))
        performance_this_case [:] = np.nan
        performance = np.vstack([performance, performance_this_case])
    else:
        performance = np.vstack([performance,evaluate_alg_CASE(model,xf,y_median, report=False)])

    return performance


def nearest_multiple(number, multiple):
    """
    Deep Learning pooling layers reduce the time-series by a factor of "multiple".
    This method ensures that the time-series windows is a multiple of the factor, and that 
    avoids runnings errors. 

    returns: nearest multiple of "number" (integer). 

    """

    nearest = round(number / multiple) * multiple  # nearest_multiple

    print("Nearest multiple of", multiple, "to", number, "is:", nearest)

    return nearest


# -----------------------------------------------------------------------
# -----------------------------------------------------------------------
# -----------------------------------------------------------------------
# -----------------------------------------------------------------------
# -----------------------------------------------------------------------
# DEAP and K-EmoCon processing functions

def extract_data_xls(sheet, NUMBER_LABELS,NUMBER_SAMPLES_PER_LABEL):
    """
    # Extract signal chunks in a per label basis
    # Returned shape is (NUMBER_LABELS, 1, PPG_NUMBER_SAMPLES_PER_LABEL) 
    """
    data = np.zeros((NUMBER_LABELS,1,NUMBER_SAMPLES_PER_LABEL))

    for label in range(0,NUMBER_LABELS):
        segment_data = []  #chunks of NUMBER_SAMPLES_PER_LABEL data. 
        for row in sheet.iter_rows(min_row=2+NUMBER_SAMPLES_PER_LABEL*label, max_row=1+NUMBER_SAMPLES_PER_LABEL*(label+1), min_col=3, max_col=3, values_only=True):
            segment_data.append(row[0])
        
        data[label][0] = np.copy(segment_data)
    
    print(np.shape(data)) #without outlier detection
    return data

def extract_labels_xls(sheet, NUMBER_LABELS):
    """
    # Extract signal chunks in a per label basis
    # Returned shape is (NUMBER_LABELS, 1, 2).
    # Last column order: arousal, valence 
    """
    labels = np.zeros((NUMBER_LABELS,2))

    arousal_labels = []
    for row in sheet.iter_rows(min_row=2, max_row=1+NUMBER_LABELS, min_col=2, max_col=2, values_only=True):
        arousal_labels.append(row[0])

    arousal_labels = np.reshape(arousal_labels, (NUMBER_LABELS,1))

    valence_labels = []
    for row in sheet.iter_rows(min_row=2, max_row=1+NUMBER_LABELS, min_col=3, max_col=3, values_only=True):
        valence_labels.append(row[0])

    valence_labels = np.reshape(valence_labels, (NUMBER_LABELS,1))

    labels = np.hstack((arousal_labels, valence_labels))    

    return labels


def plot_freq_response(FS, w, h, title):
    "Utility function to plot response functions"
    fig = plt.figure()
    ax = fig.add_subplot(111)
    ax.plot(0.5*FS*w/np.pi, 20*np.log10(np.abs(h)))
    ax.set_ylim(-40, 5)
    ax.set_xlim(0, 0.5*FS)
    ax.grid(True)
    ax.set_xlabel('Frequency (Hz)')
    ax.set_ylabel('Gain (dB)')
    ax.set_title(title)
    plt.show(block=True)

def preprocessGSR (GSR,jitterflag=0):
    """
    1) Baseline drift removal. De-trending: from neurokit2 import signal_detrend. 
    2) Median filter: median GSR score of the +/- 4 sec surrounding samples to obtain phasic component
    2.1) Jittering: add jitter to signal (if jitterflag = 1)
    2) Band pass filtering to better obtain phasic component. 
    3) Scaling: based on first 10 sec, to ensure all signal segments goes from -1 to 1. This allow inter-participant comparison. 
    4) Based on preprocessgsrbck3.py and preprocessgsrbck4.py. 
    
    If jitterflag = 1 then jittering is performed on the signal. 
    """
       
    def jitter(x, sigma=0.03):
        # See Terry Um paper
           
        # sigma = np.std(x)/40  #50
        return x + np.random.normal(loc=0., scale=sigma, size=x.shape)
    
    
    N = len(GSR)
    # print("GSR sample length: ", N)
    T = 1.0 / GSR_FS # sample spacing
    t = np.linspace(0, N*T, N, endpoint=False)
    FS = GSR_FS
       

    GSRo = np.copy(GSR)
    # detrend signal
    GSRd = signal_detrend(GSR, method='polynomial',order=3)
       
    # MEDIAN filter of the +/- 4 sec surrounding samples to obtain phasic component
    # wlength = 60 # windows length in seconds
    
    GSR = GSRd # don't slice the signal yet, to avoid median filter zero padding effects.   
    GSR  = GSR - GSR[0]  #this ensures the median filter zero padding doesn't modify the phasic signal.

    GSRfilt=medfilt(GSR,kernel_size = int(FS)*4+1)  #median GSR score of the +/- 4 surrounding samples | +1 to ensure odd kernel size. See Galvanic Skin Response – The Complete Pocket Guide 
    #remove tonic component:
    GSRphasic = GSR - GSRfilt #60 sec long
 #    GSRphasicw = GSRphasic[0:wdurationGSR] # "wlength" sec long | without median filter zero padding effect on both borders. 
       
    if jitterflag:
         ms = sum(np.square(GSRphasic))/len(GSRphasic)  #mean-square of the signal
         sigma = np.sqrt(ms/10**(SNR/10))  # SNR = 10*log10(E(S^2)/sigma^2)  # http://www.scholarpedia.org/article/Signal-to-noise_ratio#:~:text=Signal%2Dto%2Dnoise%20ratio%20is%20also%20defined%20for%20random%20variables,2_N%20the%20variance%20of%20N%5C%20.
         GSRphasic = jitter(GSRphasic,sigma) # add jitter to signal, j = jitter
         
    
    # bandpass filter GSR signal:
    
    # Heartpy filter_signal function can be used also in GSR signals (see its definition: scipy functions are used)
    ysbrbpfiltj = hp.filter_signal(GSRphasic, [LOW_CUT_GSR, HIGH_CUT_GSR], sample_rate = FS, order = ORDER_GSR_FILTER, filtertype='bandpass')  #br = baseline removed, bpfilt = bandpass filtered, j = jitter
    
    # scaling
 #    ysbrbpfiltjw = ysbrbpfiltj[0:wdurationGSR] # w = windowed
 #    ysbrbpfiltw = ysbrbpfiltj 
    ysbrbpfiltjs = hp.scale_data(ysbrbpfiltj,lower=-1,upper=1)  # s = scaled
        
#      # check signal:
#     plt.ion()
#     plt.figure()
#  #    plt.clf()

#     plt.plot(t, GSRo, label='Noisy signal')
#     plt.plot(t,GSRd, label='detrended signal')   
#     plt.plot(t, GSR, label='detrended signal - GSRd(0)')
#     plt.plot(t,GSRfilt, label='median filtered signal detr-GSR')
#     plt.plot(t,GSRphasic, label='Phasic signal')
#     plt.grid()
#     plt.legend()
#     plt.show(block=True)

    # plt.plot(t,GSRphasic, label='Phasic signal')
    # plt.grid()
    # plt.legend()
    # plt.show(block=True)


 #    plt.clf()
 #    plt.plot(t[0:len(GSRphasic)],ysbrbpfiltjs, label='+jitter')
 #    plt.grid()
 #    plt.legend()
 #    plt.show(block=True)
    
    return ysbrbpfiltjs

def preprocessPPG(PPG, jitterflag=0):
    """
    1. Baselined and DC component is removed using notch filter al 0.001 Hz (using heartpy. hp.remove_baseline_wander) and subtraction of the mean. 
    1.1 Jitter is added to signal once the baseline is removed (if jitterflag = 1)
    2. Motion artifact is removed with band pass filter – 0.05 to 18 Hz, order = 5 . 
    3. Signal is scaled (between -1 and 1).
    4. Median filter NOT applied because it blurs the the exact location of systole and diastole. 
    
    If jitterflag = 1 then jittering is performed on the signal. 
    """
        
    
    #  PPG measurement window duration is 60 seconds (see DEAP dataset markers - timestamps, One Note 2016). Windows duration greater than GSR windows because HRV needs more data... | see One Note 2016: DEAP dataset markers - timestamps
    
    N = len(PPG)
    T = 1.0 / PPG_FS # sample spacing
    t = np.linspace(0, N*T, N, endpoint=False)
    FS = PPG_FS
    
    def jitter(x, sigma=0.03):
        # See Terry Um paper
            
        # sigma = np.std(x)/40  #50
        return x + np.random.normal(loc=0., scale=sigma, size=x.shape)
        
    
    # remove baseline wander:
    xbaseremov = hp.remove_baseline_wander(PPG,FS,cutoff=0.001)
    
    if jitterflag:
        ms = sum(np.square(xbaseremov))/len(xbaseremov)  #mean-square of the signal
        sigma = np.sqrt(ms/10**(SNR/10))  # SNR = 10*log10(E(S^2)/sigma^2)  # http://www.scholarpedia.org/article/Signal-to-noise_ratio#:~:text=Signal%2Dto%2Dnoise%20ratio%20is%20also%20defined%20for%20random%20variables,2_N%20the%20variance%20of%20N%5C%20.
        xbaseremov = jitter(xbaseremov,sigma) # add jitter to signal, j = jitter
    
    # bandpass filter PPG signal:
    ybrbpfiltj = hp.filter_signal(xbaseremov, [LOW_CUT_PPG, HIGH_CUT_PPG], sample_rate = FS, order = ORDER_PPG_FILTER, filtertype='bandpass')  #br = baseline removed, bpfilt = bandpass filtered , , j = jitter
        
    #scaling:
    ybrbpfiltsj = hp.scale_data(ybrbpfiltj,lower=-1,upper=1)  # s = scaled, j = jitter
    
    # plt.ion()
    # plt.figure()
    # plt.clf()
    # plt.grid()
    # plt.plot(t, PPG, label='noisy signal')
    # plt.plot(t, ybrbpfiltsj, label='pre-processed signal')
    # plt.legend(loc='best')
    # plt.show(block=True)

    # plt.ion()
    # plt.figure()
    # plt.clf()
    # plt.grid()
    # plt.plot(t, ybrbpfiltsj, label='pre-processed signal')
    # plt.legend(loc='best')
    # plt.show(block=True)
    
    return ybrbpfiltsj    
    
    


    
def preprocessGSR_DL (GSR,jitterflag=0, phaseflag=0):
    """
    1) Baseline drift removal. De-trending: from neurokit2 import signal_detrend. 
    2) Median filter: median GSR score of the +/- 4 sec surrounding samples to obtain phasic component
    2.1) Jittering: add jitter to signal (if jitterflag = 1)
    2) Band pass filtering to better obtain phasic component. 
    3) Scaling: based on first 10 sec, to ensure all signal segments goes from -1 to 1. This allow inter-participant comparison. 
    4) Based on preprocessgsrbck3.py and preprocessgsrbck4.py. 

    If jitterflag = 1 then jittering is performed on the signal. 
    If phaseflag = 1 then phase augmentation (bootstrapp) is performed on the signal. 
    """
   
    def jitter(x, sigma=0.03):
        # See Terry Um paper
       
        # sigma = np.std(x)/40  #50
        return x + np.random.normal(loc=0., scale=sigma, size=x.shape)


    N = len(GSR)
    T = 1.0 / GSR_FS # sample spacing
    t = np.linspace(0, N*T, N, endpoint=False)
    FS = GSR_FS

    # detrend signal
#     GSRd = signal_detrend(GSR, method='polynomial',order=3)
   
#     # MEDIAN filter of the +/- 4 sec surrounding samples to obtain phasic component
#     # wlength = 60 # windows length in seconds
#     GSR = GSRd # don't slice the signal yet, to avoid median filter zero padding effects.   
#     GSR  = GSR - GSR[0]  #this ensures the median filter zero padding doesn't modify the phasic signal.
   
#     GSRfilt=medfilt(GSR,kernel_size = int(FS)*8+1)  #median GSR score of the +/- 4 sec surrounding samples | +1 to ensure odd kernel size. See Galvanic Skin Response – The Complete Pocket Guide 
#     #remove tonic component:
#     GSRphasic = GSR - GSRfilt #60 sec long
   
#     if jitterflag:
#          ms = sum(np.square(GSRphasic))/len(GSRphasic)  #mean-square of the signal
#          sigma = np.sqrt(ms/10**(SNR/10))  # SNR = 10*log10(E(S^2)/sigma^2)  # http://www.scholarpedia.org/article/Signal-to-noise_ratio#:~:text=Signal%2Dto%2Dnoise%20ratio%20is%20also%20defined%20for%20random%20variables,2_N%20the%20variance%20of%20N%5C%20.
#          GSRphasic = jitter(GSRphasic,sigma) # add jitter to signal, j = jitter
     
#     if phaseflag:
#         fft1 = fft(GSRphasic)
#         fft2 = fft(GSRphasic)
#         signal_lenght=len(GSRphasic)
#         power2=np.zeros(signal_lenght)
#         fase2=np.zeros(signal_lenght)
#         for i in np.arange(0, signal_lenght):
#             power2[i]=np.abs(fft2[i])
#             fase2[i]=np.angle(fft1[i])
       
       
#         # Tres métodos para variar la fase, elegir el que mas te gusta
#      #    np.random.shuffle(fase2)
#      #    fase2=fase2*np.random.uniform(low=0, high=.1)
#         for i in np.arange(0, signal_lenght):
#             fase2[i]=np.angle(fft2[i])*np.random.uniform(low = phase_low_threshold, high = phase_high_threshold)
       
#         for i in np.arange(0, signal_lenght):
#             if i<(int(signal_lenght/2)+1):
#                 fft2[i]=power2[i]*np.cos(fase2[i])+power2[i]*np.sin(fase2[i])*1j    
#             if i>(((signal_lenght/2))+1):
#                 fft2[i]=fft2.real[signal_lenght-(i)]-fft2.imag[signal_lenght-(i)]*1j
       
#         """
#         fft2pos=np.zeros(int(len(S1)/2))
#         fft2pos = fft2[range(int(len(S1)/2))]
#         """
       
#      #    GSRphasicp=ifft(fft2)
#         GSRphasic=np.real(ifft(fft2))
   

#     # bandpass filter GSR signal:

#     # Heartpy filter_signal function can be used also in GSR signals (see its definition: scipy functions are used)
#     ysbrbpfiltj = hp.filter_signal(GSRphasic, [LOW_CUT_GSR, HIGH_CUT_GSR], sample_rate = FS, order = ORDER_GSR_FILTER, filtertype='bandpass')  #br = baseline removed, bpfilt = bandpass filtered, j = jitter
#  #    if phaseflag:
#  #        ysbrbpfiltp = hp.filter_signal(GSRphasicp, [LOW_CUT_GSR, HIGH_CUT_GSR], sample_rate = FS, order = ORDER_GSR_FILTER, filtertype='bandpass')  #br = baseline removed, bpfilt = bandpass filtered, p = phase augmentation

    # scaling
 #    ysbrbpfiltw = ysbrbpfiltj 
    # ysbrbpfiltjs = hp.scale_data(ysbrbpfiltj,lower=-1,upper=1)  # s = scaled

    ysbrbpfiltjs = hp.scale_data(GSR,lower=-1,upper=1)  # s = scaled
    
     # check signal:
     # if (show_plots):
     #     plt.ion()
     #     plt.figure()
     #     #    plt.clf()
     #     plt.plot(t, GSR, label='Noisy signal')
        
     #     plt.plot(t,GSRfilt, label='median filtered signal')
     #     plt.grid()
     #     plt.legend()
     #     plt.show(block=True)
        
     #     plt.clf()
     #     plt.plot(t[0:len(GSRphasic)],ysbrbpfiltjs, label='+jitter')
     #     plt.grid()
     #     plt.legend()
     #     plt.show(block=True)

     #     if phaseflag:
     #         plt.ion()
     #         plt.figure()
     #         plt.clf()
     #         plt.grid()
     #         plt.plot(t, ysbrbpfiltp, label='phase augmentation')
     #         plt.plot(t, ysbrbpfiltj, label='original')
     #         plt.legend(loc='best')
     #         plt.show(block=True)


    return ysbrbpfiltjs


def preprocessPPG_DL(PPG, jitterflag=0, phaseflag=0):
    """
    1. Baselined and DC component is removed using notch filter al 0.001 Hz (using heartpy. hp.remove_baseline_wander) and subtraction of the mean. 
    1.1 Jitter is added to signal once the baseline is removed (if jitterflag = 1)
    2. Motion artifact is removed with band pass filter – 0.05 to 18 Hz, order = 5 . 
    3. Signal is scaled (between -1 and 1).
    4. Median filter NOT applied because it blurs the the exact location of systole and diastole. 

    If jitterflag = 1 then jittering is performed on the signal. 
    If phaseflag = 1 then phase augmentation (bootstrapp) is performed on the signal. 
    """
    

    #  PPG measurement window duration is 60 seconds (see DEAP dataset markers - timestamps, One Note 2016). Windows duration greater than GSR windows because HRV needs more data... | see One Note 2016: DEAP dataset markers - timestamps

    N = len(PPG)
    T = 1.0 / PPG_FS # sample spacing
    t = np.linspace(0, N*T, N, endpoint=False)
    FS = PPG_FS

    # def jitter(x, sigma=0.03):
    #     # See Terry Um paper
        
    #     # sigma = np.std(x)/40  #50
    #     return x + np.random.normal(loc=0., scale=sigma, size=x.shape)
    

    # # remove baseline wander:
    # xbaseremov = hp.remove_baseline_wander(PPG,FS,cutoff=0.001)

    # if jitterflag:
    #     ms = sum(np.square(xbaseremov))/len(xbaseremov)  #mean-square of the signal
    #     sigma = np.sqrt(ms/10**(SNR/10))  # SNR = 10*log10(E(S^2)/sigma^2)  # http://www.scholarpedia.org/article/Signal-to-noise_ratio#:~:text=Signal%2Dto%2Dnoise%20ratio%20is%20also%20defined%20for%20random%20variables,2_N%20the%20variance%20of%20N%5C%20.
    #     xbaseremov = jitter(xbaseremov,sigma) # add jitter to signal, j = jitter

    # if phaseflag:
    #     fft1 = fft(xbaseremov)
    #     fft2 = fft(xbaseremov)
    #     signal_lenght=len(xbaseremov)
    #     power2=np.zeros(signal_lenght)
    #     fase2=np.zeros(signal_lenght)
    #     for i in np.arange(0, signal_lenght):
    #         power2[i]=np.abs(fft2[i])
    #         fase2[i]=np.angle(fft1[i])
        
        
    #     # Tres métodos para variar la fase, elegir el que mas te gusta
    #     # np.random.shuffle(fase2)
    #     # fase2=fase2*np.random.uniform(low=0, high=0.1)
    #     for i in np.arange(0, signal_lenght):
    #         fase2[i]=np.angle(fft2[i])*np.random.uniform(low = phase_low_threshold, high = phase_high_threshold)
        
    #     for i in np.arange(0, signal_lenght):
    #         if i<(int(signal_lenght/2)+1):
    #             fft2[i]=power2[i]*np.cos(fase2[i])+power2[i]*np.sin(fase2[i])*1j    
    #         if i>(((signal_lenght/2))+1):
    #             fft2[i]=fft2.real[signal_lenght-(i)]-fft2.imag[signal_lenght-(i)]*1j
        
    #     """
    #     fft2pos=np.zeros(int(signal_lenght/2))
    #     fft2pos = fft2[range(int(signal_lenght/2))]
    #     """
        
    #     # xbaseremov2=ifft(fft2)
    #     xbaseremov=np.real(ifft(fft2))

    #     # if  (show_plots):
    #     #     plt.ion()
    #     #     plt.figure()
    #     #     plt.clf()
    #     #     plt.grid()
    #     #     plt.plot(t, xbaseremov2, label='phase augmentation')
    #     #     plt.plot(t, xbaseremov, label='original')
    #     #     plt.legend(loc='best')
    #     #     plt.show(block=True)

    

    # # bandpass filter PPG signal:
    # ybrbpfiltj = hp.filter_signal(xbaseremov, [LOW_CUT_PPG, HIGH_CUT_PPG], sample_rate = FS, order = ORDER_PPG_FILTER, filtertype='bandpass')  #br = baseline removed, bpfilt = bandpass filtered , , j = jitter
    # # if phaseflag:
    # #     ybrbpfiltp = hp.filter_signal(xbaseremov2, [LOW_CUT_PPG, HIGH_CUT_PPG], sample_rate = FS, order = ORDER_PPG_FILTER, filtertype='bandpass')  #br = baseline removed, bpfilt = bandpass filtered , , p = phase augmentation
    
    # #scaling:
    # ybrbpfiltsj = hp.scale_data(ybrbpfiltj,lower=-1,upper=1)  # s = scaled, j = jitter

    ybrbpfiltsj = hp.scale_data(PPG,lower=-1,upper=1)  # s = scaled, j = jitter

    # if  (show_plots):        
        # if phaseflag:
        #     plt.ion()
        #     plt.figure()
        #     plt.clf()
        #     plt.grid()
        #     plt.plot(t, ybrbpfiltp, label='phase augmentation')
        #     plt.plot(t, ybrbpfiltj, label='original')
        #     plt.legend(loc='best')
        #     plt.show(block=True)


        # plt.ion()
        # plt.figure()
        # plt.clf()
        # plt.grid()
        # plt.plot(t, ybrbpfiltsj, label='pre-processed signal')
        # plt.legend(loc='best')
        # plt.show(block=True)

    return ybrbpfiltsj    


def preprocessEEG (EEG,jitterflag=0):
   """
   1) Bandpass filtering to obtain theta, alpha and beta waves from EEG data.
   Theta = 4 - 8   Hz - Theta Band
   Alpha = 8 - 12   Hz - Alpha Band
   Beta = 12 - 30   Hz - Beta Band

   2) Remove first N-1 samples and keep the good part of the filtered signal. The first N-1 samples are corrupted by the initial conditions. 
   
   3) Resulting signal is N-1 samples shorter than the original. 

   4) Based on FIRfilterEEG.py
    
   If jitterflag = 1 then jittering is performed on the signal. 
   """

   def jitter(x, sigma=0.03):
       # See Terry Um paper
           
       # sigma = np.std(x)/40  #50
       return x + np.random.normal(loc=0., scale=sigma, size=x.shape)
    
   def extract_wave (EEG,band):
    # Extract this EEG waves: 
    # Theta = [4, 8]  # Hz - Theta Band
    # Alpha = [8, 12]  # Hz - Alpha Band
    # Beta = [12, 30]  # Hz - Beta Band
    
    # It applies Parks-Mc Clellan algorithm to finetune Chebyshev FIR filter 
    # on a EEG signal

    FS = 128.0         # Sample rate, Hz
    # Theta = [4, 8]  # Hz - Theta Band
    # Alpha = [8, 12]  # Hz - Alpha Band
    # Beta = [12, 30]  # Hz - Beta Band
    # band = band #  # Desired pass band, Hz
    trans_width = 1    # Width of transition from pass band to stop band, Hz
    numtaps = 100        # Size (order) of the FIR filter.
    edges = [0, band[0] - trans_width, band[0], band[1],
             band[1] + trans_width, 0.5*FS]
    taps = signal.remez(numtaps, edges, [0.1, 1.6, 0], Hz=FS) # numerator coefficients of the filter
    w, h = signal.freqz(taps, [1], worN=2000)
    if show_plots:
     plot_freq_response(FS, w, h, "Band-pass Filter")

    filtered_x = lfilter(taps, [1], EEG)

    #------------------------------------------------
    # Plot the original and filtered signals.
    #------------------------------------------------

    # The phase delay of the filtered signal.
    delay = 0.5 * (numtaps-1) / FS

    # Over the passband frequency range for a linear-phase, S-tap FIR filter,
    # group delay has been shown to be given by G = D * ts / 2 seconds,
    # where D = S–1 is the number of unit-delay elements in the filter’s delay line,
    # and ts is the sample period (1/f). See Lyons, Richard G - Understanding digital signal processing-Prentice Hall (2011), p. 212.
    # https://docs.scipy.org/doc/scipy/reference/generated/scipy.signal.group_delay.html#id1
    # https://scipy-cookbook.readthedocs.io/items/FIRFilter.html

    if show_plots:
     plt.figure()
     # Plot the original signal.
     plt.plot(t, EEG, label="original")
     # Plot the filtered signal, shifted to compensate for the phase delay.
     plt.plot(t - delay,filtered_x, 'r-', label="filtered")
     # Plot just the "good" part of the filtered signal.  The first N-1
     # samples are "corrupted" by the initial conditions.
     # plot(t[N-1:]-delay, filtered_x[N-1:], 'g', linewidth=4)
     plt.legend()
     plt.xlabel('t')
     plt.grid(True)
     plt.show(block=True)


    # FFT con EEG raw data 

     plt.figure()
     fx=fftfreq(N,T)
     yf=fft(EEG)

     plt.plot(fx,np.abs(yf), label='FFT on raw signal')
     plt.plot(fx,np.abs(fft(filtered_x)), label='FFT on filtered signal')
     plt.grid()
     plt.legend()
     plt.title('FFT')
     plt.xlabel('f (Hz)')
     plt.show(block=True)


     ws, gd = signal.group_delay((taps, [1]),FS=FS)
     plt.title('Digital filter group delay')
     plt.plot(ws, gd)
     plt.ylabel('Group delay [samples]')
     plt.xlabel('Frequency [rad/sample]')
     plt.show(block=True)

    return filtered_x[int(np.round(0.5 * (numtaps-1))):]  # Remove first N-1 samples and keep the good part of the filtered signal. The first N-1 samples are corrupted by the initial conditions. 
    
   N = len(EEG)
   T = 1.0 / 128.0 # sample spacing
   t = np.linspace(0, N*T, N, endpoint=False)

   delta = [1, 4]  # Hz - Delta Band
   theta = [4, 8]  # Hz - Theta Band
   alpha = [8, 12]  # Hz - Alpha Band
   beta = [12, 30]  # Hz - Beta Band
    
   theta_wave = extract_wave(EEG,theta)
   alpha_wave = extract_wave(EEG,alpha)
   beta_wave = extract_wave(EEG,beta)
    
   return theta_wave, alpha_wave, beta_wave


#%% -----------------------------
# ----- FEATURE EXTRACTION STAGE (GSR + PGG + EEG) ----------
# See Relevant Features in GSR and PPG (DEAP) - Godin et al (2015), in One Note 2016
    
def feature_extract_GSR_PPG (GSRphasic, PPGfilt, ppg_fs, gsr_fs):
        
    ts_ppg = 1/ppg_fs # ppg sample spacing
    ts_gsr = 1/gsr_fs # gsr sample spacing
    
    # GSR feature extraction:
    # feature: Average of the derivative  
    derivative=np.gradient(GSRphasic,ts_gsr)  #take the derivative (phasic component)
    deraveg=np.mean(derivative)  #feature: Average of the derivative
        
    # feature: % of neg. samples in the derivative. 
    negsamples = [num for num in derivative if num < 0]  #list comprehension
    perc_negsamples = len(negsamples)/len(derivative) #feature: % of neg. samples in the derivative. 
        
    #feature: number of local minima
    peaks, _ = find_peaks(-GSRphasic,width=10,distance=80) #negate xphasic to have find_peaks function find local minima
    num_minima=len(peaks) #feature: number of local minima
    
    # PPG feature extraction:
    wd, m = hp.process(PPGfilt, sample_rate = ppg_fs, calc_freq=True) #signal is 1 minute, two short for freq domain measure computation...
        
    #m.keys() print dict_keys
    # BPM = beats per minute (BPM) 
    # IBI = Mean inter-beat interval (IBI). See https://openresearchsoftware.metajnl.com/articles/10.5334/jors.241/
    # SDNN - (6) standard deviation of heart rate, standard deviation if intervals between adjacent beats, SDNN
    # RMSSD- (7) Root Mean Square of Successive Differences (RMSSD), root mean square of successive differences between adjacend R-R intervals, RMSSD
    # SDSD - (8) Standard Deviation of Successive Differences (SDSD), standard deviation of successive differences between adjacent R-R intervals, SDSD
    # HF - High-frequency band - HRV can also be expressed in the frequency domain
        
        
    # //////////
    # MIN-MAX features VALUES:
        
    # # BPM = beats per minute (BPM) 
    # 40-200 bpm...  https://www.heart.org/en/healthy-living/fitness/fitness-basics/target-heart-rates
        
    # # IBI = Mean inter-beat interval (IBI). See https://openresearchsoftware.metajnl.com/articles/10.5334/jors.241/
    # [ms] 500-1300 ms  || http://www.anslab.net/static/help/cardiography.html
        
    # # SDNN - (6) standard deviation of heart rate, standard deviation if intervals between adjacent beats, SDNN - Standard deviation of NN intervals
    # 16-110 ms ||| see Van den Berg, 2018
        
    # # RMSSD- (7) Root Mean Square of Successive Differences (RMSSD), root mean square of successive differences between adjacend R-R intervals, RMSSD
    # 17-170 ms ||| see Van den Berg, 2018
        
    # # SDSD - (8) Standard Deviation of Successive Differences (SDSD), standard deviation of successive differences between adjacent R-R intervals, SDSD
    # 10-100 ms ||| see Cui, 2020        
        
    # # HF - High-frequency band - Absolute power of the high-frequency band (0.15–0.4 Hz) 
    # ms^2  || 500-5500 (y más amplio también) || see Cuid, 2020 y https://www.intellihinc.com/heart-rate-variability/
        
        

    # ATENTION: freq domain features: they are not very reliable given the raw signal is only 1 minute long.
    # See Shaffer et al.
    # (9) Heart Rate Variability (HRV) power in the bands VLF (0.01-0.04] Hz, 
    # (10) LF (0.04-0.15] Hz,  which is related to short-term blood pressure variation. See https://openresearchsoftware.metajnl.com/articles/10.5334/jors.241/
    # (11) HF (0.15-0.5] Hz, which is a reflection of breathing rate. See https://openresearchsoftware.metajnl.com/articles/10.5334/jors.241/
    # HF

    # fv = [deraveg, perc_negsamples, num_minima, m['bpm'], m['ibi'], m['sdnn'], m['rmssd'], m['sdsd'], m['vlf'], m['lf'], m['hf']]
    # fv = [deraveg, perc_negsamples, num_minima, m['bpm'], m['ibi'], m['sdnn'], m['rmssd'], m['sdsd'], m['hf']]
    fv = [deraveg, perc_negsamples, num_minima, m['bpm'], m['ibi'], m['sdnn'], m['rmssd'], m['sdsd']]
    
    return fv  #feature vector
    
# print(feature_extract_GSR_PPG(GSRphasic, PPGfilt))


def featureextract_eeg (EEG):
        
    # FS = sampling freq. 

    def second_dif(x):

        N = len(x)
    
        xdiff = np.zeros((N-2))

        for i in range(0,N-2):
            xdiff[i] = x[i+2] - x[i]

        return xdiff
    
    def standarize_signal(signal):

        mean = np.mean(signal)   # mean value
        std = np.std(signal)     # standard deviation

        standarized_signal = [(x-mean)/std for x in signal]

        return standarized_signal


    mean = np.mean(EEG)   # mean value
    std = np.std(EEG)     # standard deviation
    diff = np.diff(EEG,1)
    afd = np.mean(np.absolute(diff)) # AFD - Mean of the absolute values of the first differences
    standarized = standarize_signal(EEG)
    diffs = np.diff(standarized,1) # s = standarized
    afdn = np.mean(np.absolute(diffs)) # AFD - Mean of the normalised (standarized) absolute values of the first differences
    diff2 = second_dif(EEG)    
    asd = np.mean(np.absolute(diff2)) # ASD: Mean of the absolute values of the second differences
    diff2s = second_dif(standarized)     # s = standarized
    asdn = np.mean(np.absolute(diff2s)) # ASDN - Mean of the normalised (standarized) absolute values of the second differences (ASDN)

    return [mean,std,afd,afdn,asd,asdn]  #feature vector



# https://machinelearningmastery.com/multi-class-imbalanced-classification/
# load and summarize the dataset

def class_imbalance_test (labels, show_plots):

    one_class = False  # False if dataset has only only class or one of the two classes contains only one sample
    labels_o = np.ravel(labels)
    if (len(labels_o) != 0):
        # summarize distribution
        counter = Counter(labels_o)
        print("Multiclass distribution")
        for k,v in counter.items():
            per = v / len(labels_o) * 100
            print('Class=%d, n=%d (%.3f%%)' % (k, v, per))
            if (v == 0) or (v == 1) or (per == 100.0):    
                one_class = True
        if  (show_plots):
            # plot the distribution
            plt.bar(counter.keys(), counter.values())
            plt.show()
    else: 
        one_class = True
    return one_class
    



# Metrics calculation

def calc_metrics(expected, predicted):

    # ConfusionMatrix
    #                | Positive Prediction | Negative Prediction
    # Positive Class | True Positive (TP)  | False Negative (FN)
    # Negative Class | False Positive (FP) | True Negative (TN)

    #  4    1
    #  3    22

    results = confusion_matrix(expected, predicted)
    print("confusion matrix: ")
    print(results)

    # calculate prediction
    # Precision = TruePositives / (TruePositives + FalsePositives)
    # precision = precision_score(expected, predicted, average='binary', pos_label = 1)
    # print('Precision (ones): %.3f' % precision)


    # acc = accuracy_score(expected, predicted)
    # print('Accuracy: %.3f' % acc) #same as MATLAB


    # # Recall = TruePositive / (TruePositive + FalseNegative)
    # recall = recall_score(expected, predicted, average='binary', pos_label = 1)
    # print('Recall (ones): %.3f' % recall    )


    
    # # Recall = TruePositive / (TruePositive + FalseNegative)
    # f1score = f1_score(expected, predicted, average='binary', pos_label = 1)
    # print('f1-score (ones): %.3f' % f1score)


    
    # Recall = TruePositive / (TruePositive + FalseNegative)
    class_report = classification_report(expected, predicted)
    print(class_report)

    class_report = classification_report(expected, predicted,output_dict=True)
    f1_score_macro = class_report['macro avg']['f1-score']
    print("macro avg f1-score:", f1_score_macro)
    uar = class_report['macro avg']['recall']
    print("macro avg (UAR):", uar)

    cohen_kappa_score_value = cohen_kappa_score(expected,predicted)


    # ConfusionMatrix
    #                | Positive Prediction | Negative Prediction
    # Positive Class | True Positive (TP)  | False Negative (FN)
    # Negative Class | False Positive (FP) | True Negative (TN)

    # sensitivity:

    if len(results) == 2: #if confusion matrix is 2 x 2 
        TP = results[0,0] #True Positive
        FN = results[0,1]
        FP = results[1,0]
        TN = results[1,1]

        # Sensitivity = TruePositive / (TruePositive + FalseNegative)
        # Recall, True Positive Rate
        Sensitivity = TP / (TP + FN)
        print("Sensitivity: ", Sensitivity)

        # Specificity = TrueNegative / (FalsePositive + TrueNegative)
        # True Negative rate
        Specificity = TN / (FP + TN)
        print("Specificity: ", Specificity)

        gmean = np.sqrt(Sensitivity*Specificity)
        print("g-mean: ", gmean)

    elif len(results) == 1: #if confusion matrix is 1 x 1 ()

        if int(np.mean(expected)) == 1:

            # Sensitivity = TruePositive / (TruePositive + FalseNegative)
            Sensitivity = 1   # recall of the positive class
            print("Sensitivity: ", Sensitivity)

            # Specificity = TrueNegative / (FalsePositive + TrueNegative)
            Specificity = 0   # recall of the negative class
            print("Specificity: ", Specificity)

            gmean = np.sqrt(Sensitivity*Specificity)
            print("g-mean: ", gmean)

        else: 
            # Sensitivity = TruePositive / (TruePositive + FalseNegative)
            Sensitivity = 0   # recall of the positive class
            print("Sensitivity: ", Sensitivity)

            # Specificity = TrueNegative / (FalsePositive + TrueNegative)
            Specificity = 1   # recall of the negative class
            print("Specificity: ", Specificity)

            gmean = np.sqrt(Sensitivity*Specificity)
            print("g-mean: ", gmean)
    
    # # Fbeta = ((1 + beta^2) * Precision * Recall) / (beta^2 * Precision + Recall)
    # f2_score = fbeta_score(expected, predicted, beta = 2)
    # print("f2_score: ", f2_score)

    return [cohen_kappa_score_value, uar, gmean, f1_score_macro]

# expected = [0,0,0,0,0,0]
# predicted = [0,0,0,0,0,0]
# calc_metrics(expected, predicted)


# create a feature preparation pipeline for a model
def make_pipeline(model):
	steps = list()
	# standardization
	steps.append(('standardize', StandardScaler()))
	# normalization
	steps.append(('normalize', MinMaxScaler()))
	# the model
	steps.append(('model', model))
	# create pipeline
	pipeline = Pipeline(steps=steps)
	return pipeline


# compute accuracy (sklearn)
def accuracy_calculation(model,input,expected):

    print("prediction:", model.predict(input))
    print("expected:",expected)
    acc = model.score(input,expected)
    print("Accuracy (test set): %0.2f " % acc)  # mean accuracy, fraction of correctly classified samples; https://scikit-learn.org/stable/modules/generated/sklearn.metrics.accuracy_score.html#sklearn.metrics.accuracy_score

    return f"{round(acc,2):.2f}"  # round to 2 decimals

from sklearn.base import clone

def standarize_normalize(x_train, x_test):

    # Standardize and normalize the training and test datasets independently

    # stdscaler = StandardScaler
    # minmaxcaler = MinMaxScaler

    stdscaler = StandardScaler().fit(x_train)
    x_train_std = stdscaler.transform(x_train)
    minmaxscaler = MinMaxScaler().fit(x_train_std)
    x_train_sn = minmaxscaler.transform(x_train_std)

    x_test_std = stdscaler.transform(x_test)
    x_test_sn = minmaxscaler.transform(x_test_std)

    return x_train_sn, x_test_sn


#%% -----------------------------
def split_test_alg(model,X,y):


    
    # 1) split, 2) test ALG, 3) repeat 1 and 2 (bootstrap)

    # It performs a stratified Split of dataset, it trains the model using a pipeline and
    # and evaluate it using metric calculations. 
    
    # StratifiedShuffleSplit

    # This cross-validation object returns stratified randomized folds. The folds are made by preserving the percentage of samples for each class.
    # Note: like the ShuffleSplit strategy, stratified random splits do not guarantee that all folds will be different, although this is still very likely for sizeable datasets.

    # it perfoms k-fold and ramdom split, preserving the percentage of samples for each class, if there are
    # only two clases. Useful for imbalanced datasets. 

    # Only if the number of classes is equal to one, it performs a split preserving (aproximatedly) the percentage of the 
    # minority/mayority class both  in the train and test datasets. 

    # X and y must be numpy arrays (flattened)
    
    REPETITIONS = 10  # the binomial coefficient of n or list all possible combinations of a set of items (n = 10, k = 8). Worst case: 10 labels, train ds labels= 8. Original number: 45

    # REPETITIONS = 5  # the binomial coefficient of n or list all possible combinations of a set of items (n = 10, k = 8). Worst case: 10 labels, train ds labels= 8. Original number: 45

    acc_array = np.zeros(REPETITIONS)
    gmean_array = np.zeros(REPETITIONS)
    f1_score_macro_array = np.zeros(REPETITIONS)
    uar_array = np.zeros(REPETITIONS)
    cohen_array = np.zeros(REPETITIONS)


    model_copy = clone(model) #backup of the model to retrieve the original later on. 

    print("---------------------")
    print("----- Model --------")
    print(model)
    print("---------------------")

    for i in range(0,REPETITIONS):
        
        print("--------index i----------: ", i)
        
        sss = StratifiedShuffleSplit(n_splits=1, test_size=TEST_DS_SIZE, random_state=None) # valence / arousal StratifiedShuffleSplit
        sss.get_n_splits(X, y)  #get train and test datasets. 

        print(sss)

        # The following lines are only intented to get the train and test indexes, that later
        # will be used to get the train and test dataset, with the corresponding features.   

        for train_index, test_index in sss.split(X, y):  #https://scikit-learn.org/stable/modules/generated/sklearn.model_selection.StratifiedShuffleSplit.html
            print("TRAIN:", train_index, "TEST:", test_index)
            x_train, x_test = X[train_index], X[test_index]  #_v: valence
            y_train, y_test = y[train_index], y[test_index]

        print("(SL) TRAIN number of instances: ", np.max(np.shape(y_train)))
        print("(SL) TEST number of instances: ", np.max(np.shape(y_test)))
        print("(SL) Total number of instances (TRAIN+TEST): ", np.max(np.shape(y_train))+np.max(np.shape(y_test)))

        y_train = y_train.flatten()  #arousal y_train datasets  / labels || constains augmentated data
        y_test = y_test.flatten()  #arousal y_test datasets  / labels


        [x_train, x_test] = standarize_normalize(x_train, x_test)

        # prepara data without data leakage:
        # https://machinelearningmastery.com/data-preparation-without-data-leakage/
        # "The correct approach to performing data preparation with a train-test split
        # evaluation is to fit the data preparation on the training set, then apply 
        # the transform to the train and test sets.
        # This avoids data leakage as the calculation of the minimum and maximum 
        # value for each input variable is calculated using only the training 
        # dataset (X_train) instead of the entire dataset (X)"":

        model_copy.fit(x_train,np.ravel(y_train))

        acc_array[i] = model_copy.score(x_test,y_test)
        print("acc method 1: ", acc_array[i]) #it uses the pipeline
        acc_array[i] = accuracy_calculation(model_copy,x_test,y_test) 
        # print("acc method 2", acc) #it uses the pipeline

        [cohen_array[i], uar_array[i], gmean_array[i], f1_score_macro_array[i]]  = calc_metrics(y_test, model_copy.predict(x_test))
        # [gmean_array[i], f1_score_macro_array[i]]  = calc_metrics(y_test, model_copy.predict(x_test))
        del model_copy
        model_copy = clone(model)

    print("-------- Model Performance ----------: ")    
    print("accuracy: ", acc_array)
    print("gmean: ", gmean_array)
    print("f1_score: ", f1_score_macro_array)
    print("UAR: ", uar_array)
    print("Cohen Kappa score: ", cohen_array)


    # return np.array([np.nanmean(acc_array),np.nanmean(gmean_array),np.nanmean(f1_score_macro_array)],dtype='float')
    return np.array([np.nanmean(cohen_array), np.nanmean(uar_array), np.nanmean(acc_array),np.nanmean(gmean_array),np.nanmean(f1_score_macro_array)],dtype='float')


#% Map Valence/Arousal binary and separate values to integer values (4 quadrants)
    
def map_quadrants(x):  
    if x == [1,1]:
        return 1
    elif x == [0,1]: 
        return 2    
    elif x == [0,0]: 
        return 3    
    elif x == [1,0]: 
        return 4
    
# Map Valence/Arousal values (float) to binary values (0/1)
# returns: labelsint, an array of valence and arousal. 
# medium values ( 3 < x < 7  ) are mapped to a an arbitrary high number (to be later discarted)


def map_labels(x):  #maps valence and arousal (from 1.0 to 9.0 (float) scale to a 0 to 1 binary scale)
    if x <= LOW_LABEL_THRESHOLD:
        return 0
    elif x >= HIGH_LABEL_THRESHOLD:
        return 1
    # else: 
    #     return random.choice([0, 1])
    else: 
        return MEDIUM_LABEL_VALUES  # medium partition |  medium values ( 3 < x < 7  ) are mapped to a an arbitrary high number (to be later discarted)

def evaluate_alg(model,X_a,y_a,X_v,y_v):
    #evaluate model (arousal and valence separatedly): accuraccy, geometric mean and macro f1 score


    [aro_cohen, aro_uar, aro_acc, aro_gm, aro_f1] = split_test_alg(model,X_a,y_a)
    [val_cohen, val_uar, val_acc, val_gm, val_f1] = split_test_alg(model,X_v,y_v)

    # return np.array([(val_acc,val_gm, val_f1, aro_acc,aro_gm, aro_f1)])
    return np.array([(val_cohen, val_uar, val_acc,val_gm, val_f1, aro_cohen, aro_uar,aro_acc,aro_gm, aro_f1)])

def evaluate_alg_arousal(model,X_a,y_a):
    #evaluate model (arousal only): accuraccy, geometric mean and macro f1 score

    [aro_cohen, aro_uar, aro_acc, aro_gm, aro_f1] = split_test_alg(model,X_a,y_a)

    return np.array([(np.nan, np.nan, np.nan,np.nan, np.nan, aro_cohen, aro_uar,aro_acc,aro_gm, aro_f1)]) # valence metrics =  np.nan

def evaluate_alg_valence(model,X_v,y_v):
    #evaluate model (valence only): accuraccy, geometric mean and macro f1 score

    [val_cohen, val_uar, val_acc, val_gm, val_f1] = split_test_alg(model,X_v,y_v)

    return np.array([(val_cohen, val_uar, val_acc,val_gm, val_f1, np.nan, np.nan,np.nan,np.nan, np.nan)])  # arousal metrics = np.nan
   

def split_ds(X_gsr, X_ppg,y):
    """
    Split a dataset with StratifiedShuffleSplit
    """

    from sklearn.model_selection import StratifiedShuffleSplit 
    sss = StratifiedShuffleSplit(n_splits=1, test_size=TEST_DS_SIZE, random_state=None) # valence / arousal StratifiedShuffleSplit
    sss.get_n_splits(X_gsr, y)  #get train and test datasets. 

    print(sss)

    # The following lines are only intented to get the train and test indexes, that later
    # will be used to get the train and test dataset, with the corresponding features.   

    for train_index, test_index in sss.split(X_gsr, y):  #https://scikit-learn.org/stable/modules/generated/sklearn.model_selection.StratifiedShuffleSplit.html
        print("TRAIN:", train_index, "TEST:", test_index)
        x_trainGSR, x_testGSR = X_gsr[train_index], X_gsr[test_index]  
        x_trainPPG, x_testPPG = X_ppg[train_index], X_ppg[test_index]  
        y_train, y_test = y[train_index], y[test_index]  #labels are common to GSR and PPG

    print("(DL) TRAIN number of instances: ", np.max(np.shape(y_train)))
    print("(DL) TEST number of instances: ", np.max(np.shape(y_test)))
    print("(DL) Total number of instances (TRAIN+TEST): ", np.max(np.shape(y_train))+np.max(np.shape(y_test)))

    return [x_trainGSR, x_testGSR, x_trainPPG, x_testPPG, y_train, y_test] 


def train_GSR_AE(x_trainGSR):
    """
    Train AutoEncoder for GSR signal
    Train GSR AutoEncoder, delete decoder and return encoder    
    """

    # SCAE - stacked convolutional auto-encoder 

    # Conv1D layer:
    # Input shape = 3+D tensor with shape: batch_shape + (steps, input_dim) || https://keras.io/api/layers/convolution_layers/convolution1d/
    # steps = patch_length in units of time steps
    # input_dim = input dimension of data. In this case unidimensional, because is a time series. 

    # activation = none: so autoencoder learns a signal with negative values... 
    # Thus, output of each layer is a linear combination and learns only linear patterns?

#       convolutional NN adapt for 128 Hz and a relatively high number of input samples (7000 aprox)
#     convolutional=keras.Sequential(  
#     [   
#         layers.Conv1D(filters=3, kernel_size=20, padding = "same", strides = 1, activation = "tanh", input_shape=(x_trainGSR.shape[1],x_trainGSR.shape[2]), name="1stConvL"),  
#         layers.AveragePooling1D(pool_size=4, strides=4, padding="valid", data_format="channels_last", name="1stPoolL"),        
#         layers.Conv1D(filters=3, kernel_size=10, padding = "same", strides = 1, activation = "tanh", name="2ndConvL"),  
#         layers.AveragePooling1D(pool_size=4, strides=4, padding="valid", data_format="channels_last", name="2ndPoolL"),        
#         layers.UpSampling1D(size=4),
#         layers.Conv1DTranspose(filters=3, kernel_size=10, padding = "same", strides = 1, activation = "tanh"),  
#         layers.UpSampling1D(size=4), 
#         layers.Conv1DTranspose(filters=1, kernel_size=20, padding = "same", strides = 1, activation = "tanh"),  
#     ]
# )


    convolutional=keras.Sequential(  
        [   
            layers.Conv1D(filters=5, kernel_size=20, padding = "same", strides = 1, activation = "tanh", input_shape=(x_trainGSR.shape[1],x_trainGSR.shape[2]), name="1stConvL"),  
            layers.AveragePooling1D(pool_size=4, strides=4, padding="valid", data_format="channels_last", name="1stPoolL"),        
            layers.Conv1D(filters=6, kernel_size=10, padding = "same", strides = 1, activation = "tanh", name="2ndConvL"),  
            layers.AveragePooling1D(pool_size=4, strides=4, padding="valid", data_format="channels_last", name="2ndPoolL"),        
            layers.UpSampling1D(size=4),
            layers.Conv1DTranspose(filters=6, kernel_size=10, padding = "same", strides = 1, activation = "tanh"),  
            layers.UpSampling1D(size=4), 
            layers.Conv1DTranspose(filters=1, kernel_size=20, padding = "same", strides = 1, activation = "tanh"),  
        ]
    )


    # convolutional.weights
    print("----- train_GSR_AE -------")
    convolutional.summary()
    if (show_plots):    
        keras.utils.plot_model(convolutional, 'my_sequential.png', show_shapes=True) #plot NN architecture

    #%% ---------
    
    ## Train the model

    # Please note that we are using `x_train` as both the input and the target
    # since this is a reconstruction model.
    
    convolutional.compile(optimizer=keras.optimizers.Adam(learning_rate=0.001), loss="mse",metrics=[tf.keras.metrics.MeanSquaredError()])

    history = convolutional.fit(
        x_trainGSR,
        x_trainGSR,
        epochs=5,  # original = 50
        # epochs=3,  # original = 50
        batch_size=1,
        validation_split=0.1,  #validation data extracted from the TRAIN ds, which is already shuffled. 
        # validation_data=[x_testGSR,x_testGSR],
        # shuffle=False,  #data is already shuffled with StratifiedShuffleSplit. 
        # verbose=1,
        callbacks=[
            # keras.callbacks.EarlyStopping(monitor="val_loss", patience=5, mode="min")
            keras.callbacks.EarlyStopping(monitor="loss", patience=5, mode="min")
        ],
    )

    # Let's plot training and validation loss to see how the training went.
    
    if  (show_plots):        
        plt.plot(history.history["loss"], label="Training Loss")
        plt.plot(history.history["val_loss"], label="Validation Loss")
        plt.legend()
        plt.show()

    #%%----------------------------------------------------------------
    # compare training signal vs predicted one
    # compare on data never seen before: 
    # x_train_pred = convolutional.predict(x_testGSR)

    # if  (show_plots):        
    #     plt.plot(x_testGSR[-2],label="original")
    #     plt.plot(x_train_pred[-2],label="predicted")
    #     plt.legend()
    #     plt.show()


    # x_train_pred = convolutional.predict(x_trainGSR)
    # plt.plot(x_trainGSR[-2],label="original")
    # plt.plot(x_train_pred[-2],label="predicted")
    # plt.legend()
    # plt.show()


    #%%----------------------------------------------------------------

    # feature extractor

    feature_extractor = keras.Model(
        inputs=convolutional.inputs,
        # outputs=convolutional.get_layer(name="1stConvL").output,
        outputs=convolutional.get_layer(name="2ndPoolL").output,    
    )

    features = feature_extractor(x_trainGSR)  # features = output from top layer
    # print(features)

    #%% -------------------
    # If you flatten a tensor you can see what order it is laid out in memory:
    # print(features)
    # features_flattened = tf.reshape(features, [-1]) # A `-1` passed in the `shape` argument says "Whatever fits".
    # print(features_flattened)
    # print(features_flattened[0:9])


    #%% -------------------
    # Plot Features

    feature_npa = features[:].numpy()  #convert tensor to numpy array
    # print(feature_npa)
    print(feature_npa.shape)
    # feature_npa[:,:,0]

    if  (show_plots):        
        plt.plot(x_trainGSR[0],label="original")
        plt.figure()
        plt.plot(feature_npa[1,:,:], label='feature map 1 - 1stConvL')
        # plt.plot(feature_npa[:,:,1], label='feature map 2 - 1stConvL')
        # plt.plot(feature_npa[:,:,2], label='feature map 3 - 1stConvL')
        # plt.legend()


    #%% ----------------------------------------------------------------
    #  Delete Decoder layers (while preserving bottoms layers's weights)
    for i in range(0,4):
        convolutional.pop() #delete top layers

    # convolutional.trainable = False #freeze weights
    convolutional.trainable = True #freeze weights

    convolutional.summary()  #encoder
    if  (show_plots):
        keras.utils.plot_model(convolutional, 'my_sequential.png', show_shapes=True) #plot NN architecture


    return convolutional


def  train_PPG_AE(x_trainPPG):
    """
    Train AutoEncoder for PPG signal.
    Train PPG AutoEncoder, delete decoder and return encoder
    """
    # PPGchannel

    # SCAE - stacked convolutional auto-encoder 


    # Conv1D layer:
    # Input shape = 3+D tensor with shape: batch_shape + (steps, input_dim) || https://keras.io/api/layers/convolution_layers/convolution1d/
    # steps = patch_length in units of time steps
    # input_dim = input dimension of data. In this case unidimensional, because is a time series. 

    # activation = none: so autoencoder learns a signal with negative values... 
    # Thus, output of each layer is a linear combination and learns only linear patterns?

    # adapted for DEAP, with relatively high number of inputs (7000 aprox)
    # convolutionalPPG=keras.Sequential(  
    #     [   
        
    #         layers.Conv1D(filters=5, kernel_size=20, padding = "same", strides = 1, activation = "tanh", input_shape=(x_trainPPG.shape[1],x_trainPPG.shape[2]), name="1stConvL"),  
    #         layers.AveragePooling1D(pool_size=4, strides=4, padding="valid", data_format="channels_last", name="1stPoolL"),        
    #         layers.Conv1D(filters=6, kernel_size=10, padding = "same", strides = 1, activation = "tanh", name="2ndConvL"),  
    #         layers.AveragePooling1D(pool_size=4, strides=4, padding="valid", data_format="channels_last", name="2ndPoolL"),        
    #         layers.UpSampling1D(size=4),
    #         layers.Conv1DTranspose(filters=6, kernel_size=10, padding = "same", strides = 1, activation = "tanh"),  
    #         layers.UpSampling1D(size=4), 
    #         layers.Conv1DTranspose(filters=1, kernel_size=20, padding = "same", strides = 1, activation = "tanh"),  
    #     ]
    # )

    convolutionalPPG=keras.Sequential(  
    [   
        
        layers.Conv1D(filters=5, kernel_size=20, padding = "same", strides = 1, activation = "tanh", input_shape=(x_trainPPG.shape[1],x_trainPPG.shape[2]), name="1stConvL"),  
        layers.AveragePooling1D(pool_size=4, strides=4, padding="valid", data_format="channels_last", name="1stPoolL"),        
        layers.Conv1D(filters=6, kernel_size=10, padding = "same", strides = 1, activation = "tanh", name="2ndConvL"),  
        layers.AveragePooling1D(pool_size=4, strides=4, padding="valid", data_format="channels_last", name="2ndPoolL"),        
        layers.UpSampling1D(size=4),
        layers.Conv1DTranspose(filters=6, kernel_size=10, padding = "same", strides = 1, activation = "tanh"),  
        layers.UpSampling1D(size=4), 
        layers.Conv1DTranspose(filters=1, kernel_size=20, padding = "same", strides = 1, activation = "tanh"),  
    ]
)

    convolutionalPPG.weights

    print("----- train_PPG_AE -------")
    convolutionalPPG.summary()
    if  (show_plots):
        keras.utils.plot_model(convolutionalPPG, 'my_sequential.png', show_shapes=True) #plot NN architecture



    #%% ---------
    # """
    # ## Train the model

    # Please note that we are using `x_train` as both the input and the target
    # since this is a reconstruction model.
    # """
    convolutionalPPG.compile(optimizer=keras.optimizers.Adam(learning_rate=0.001), loss="mse", metrics=[tf.keras.metrics.MeanSquaredError()])

    historyPPG = convolutionalPPG.fit(
        x_trainPPG,
        x_trainPPG,
        epochs=5,  #original = 50
        # epochs=3,  #original = 50
        batch_size=1,
        validation_split=0.1,
        callbacks=[
            # keras.callbacks.EarlyStopping(monitor="val_loss", patience=5, mode="min")
            keras.callbacks.EarlyStopping(monitor="loss", patience=5, mode="min")
        ],
    )

    # Let's plot training and validation loss to see how the training went.

    if  (show_plots):        
        plt.plot(historyPPG.history["loss"], label="Training Loss")
        plt.plot(historyPPG.history["val_loss"], label="Validation Loss")
        plt.legend()
        plt.show()

    #%%----------------------------------------------------------------
    # compare training signal vs predicted one 
    # x_train_ppg_pred = convolutionalPPG.predict(x_testPPG)

    # if  (show_plots):        
    #     plt.plot(x_testPPG[3],label="original")
    #     plt.plot(x_train_ppg_pred[3],label="predicted")
    #     plt.legend()
    #     plt.show()

    # compare training signal vs predicted one 

    # x_train_ppg_pred = convolutionalPPG.predict(x_trainPPG)
    # plt.plot(x_trainPPG[0],label="PPG original")
    # plt.plot(x_train_ppg_pred[0],label="PPG predicted")
    # plt.legend()
    # plt.show()


    #%%----------------------------------------------------------------

    # feature extractor

    feature_extractorPPG = keras.Model(
        inputs=convolutionalPPG.inputs,
        outputs=convolutionalPPG.get_layer(name="1stConvL").output,
    )

    featuresPPG = feature_extractorPPG(x_trainPPG[0])  # features = output from top layer
    # print(featuresPPG)

    # res = tf.math.count_nonzero(tf.greater_equal(featuresPPG, 0.))
    # print(res)


    #%% -------------------
    # If you flatten a tensor you can see what order it is laid out in memory:
    # print(featuresPPG)
    # features_flattened = tf.reshape(featuresPPG, [-1]) # A `-1` passed in the `shape` argument says "Whatever fits".
    # print(features_flattened)
    # print(features_flattened[0:9])


    #%% -------------------
    # Plot Features

    feature_npa = featuresPPG[:].numpy()  #convert tensor to numpy array
    # print(feature_npa)
    print(feature_npa.shape)
    # feature_npa[:,:,0]

    if  (show_plots):        
        plt.plot(x_trainPPG[0],label="original")
        plt.plot(feature_npa[:,:,0], label='feature map 1 - 1stConvL')
        plt.plot(feature_npa[:,:,1], label='feature map 2 - 1stConvL')
        plt.plot(feature_npa[:,:,2], label='feature map 3 - 1stConvL')
        plt.legend()
        plt.show()


    #%% ----------------------------------------------------------------
    # Delete Decoder layers (while preserving bottoms layers's weights)
    for i in range(0,4):
        convolutionalPPG.pop() #delete top layer

    # convolutionalPPG.trainable = False #freeze weights
    convolutionalPPG.trainable = True #freeze weights


    convolutionalPPG.summary()  #encoder
    if  (show_plots):
        keras.utils.plot_model(convolutionalPPG, 'my_sequential.png', show_shapes=True) #plot NN architecture

    return convolutionalPPG


def arousal_bimodal_deep_denoising_AE(convolutionalGSR_a, convolutionalPPG_a, ax_trainGSR, ax_testGSR, ax_trainPPG, ax_testPPG, ay_train, ay_test): 

    inputsGSR = keras.Input(shape=(ax_trainGSR.shape[1],ax_trainGSR.shape[2]),name="inputGSR")
    inputsPPG = keras.Input(shape=(ax_trainPPG.shape[1],ax_trainPPG.shape[2]),name="inputPPG")    

    x = convolutionalGSR_a(inputsGSR) # weights were freezed
    xppg = convolutionalPPG_a(inputsPPG) # weights were freezed
    y = layers.concatenate([x,xppg])
    xr = layers.Permute((2,1))(y)  #to concatenate one map after another.
    yf = layers.Flatten()(xr)
    d1 = layers.Dropout(rate=0.7)(yf)
    output = layers.Dense(1,activation="sigmoid",kernel_initializer=tf.keras.initializers.GlorotNormal(seed=12345),bias_initializer=initializers.Zeros())(d1)


    arousal_model=keras.Model(
                inputs=[inputsGSR,inputsPPG],
                # inputs=[inputsGSR],
                outputs=output, 
                name="arousal_NN"
            )    

    arousal_model.compile(optimizer=keras.optimizers.Adam(learning_rate=1e-4), loss="binary_crossentropy", metrics=[tf.keras.metrics.BinaryAccuracy()])        

    print("------------------model.summary-------------")
    print("------------------model.summary-------------")
    print("------------------model.summary-------------")
    print("------------------model.summary-------------")
    print("------------------model.summary-------------")
    print("------------------model.summary-------------")

    arousal_model.summary()

    print("------------------model.summary-------------")
    print("------------------model.summary-------------")
    print("------------------model.summary-------------")
    print("------------------model.summary-------------")
    print("------------------model.summary-------------")
    print("------------------model.summary-------------")
    

    history_a = arousal_model.fit(
        [ax_trainGSR,ax_trainPPG],  # train model with the whole TRAIN ds. 
        # [x_trainGSR],  # train model with the whole TRAIN ds. 
        [ay_train],  
        epochs=10,  #6 | Number of epochs to train the model. An epoch is an iteration over the entire x and y data provided.
        # epochs=1,  #Number of epochs to train the model. An epoch is an iteration over the entire x and y data provided.
        batch_size=1,
        # validation_split = 0.3,
        # validation_data=([x_testGSR],ay_test),
        # validation_data=([x_testGSR,x_testPPG],ay_test),
        # validation_data=([x_testGSR],ay_test), # Data on which to evaluate the loss and any model metrics at the end of each epoch. The model will not be trained on this data. 
        shuffle=True, # Boolean (whether to shuffle the training data before each epoch). A kink of evaluation using cv....
        # callbacks=[
        #     keras.callbacks.EarlyStopping(monitor="val_loss", patience=5, mode="min")
        # ],
    )

    if  (show_plots):        
        plt.plot(history_a.history["loss"], label="Training Loss")
        # plt.plot(history_a.history["val_loss"], label="Validation Loss")
        # plt.plot(history_a.history["val_binary_accuracy"], label="Validation Acc")
        plt.legend()
        plt.show()

    # compare prediction accuracy between validation accuracy (on train ds) and test acc (on TEST ds) 
    # compare training signal vs predicted one 
    # test model on data never seen before:
    prediction = arousal_model.predict([ax_testGSR,ax_testPPG])
    # prediction = nn.predict([x_testGSR])

    print("predicted", np.ndarray.flatten(prediction))
    prediction = np.ndarray.round(np.ndarray.flatten(prediction))
    prediction = prediction.astype(int)
    prediction = np.ravel(prediction)
    print("predicted", prediction)
    print("expected", ay_test)

    # aro_acc = accuracy_calculation(arousal_model,prediction,ay_test)
    # print("accuracy_calculation:", aro_acc)

    aro_acc = accuracy_score(ay_test, prediction)

    # aro_acc = np.sum(~np.logical_xor(prediction,ay_test))/len(prediction)
    print("accuracy:", aro_acc)

    [aro_cohen, aro_uar, aro_gm, aro_f1] = calc_metrics(ay_test, prediction)
    
    return [aro_cohen, aro_uar, aro_acc, aro_gm, aro_f1]


def valence_bimodal_deep_denoising_AE(convolutionalGSR_v, convolutionalPPG_v, vx_trainGSR, vx_testGSR, vx_trainPPG, vx_testPPG, vy_train, vy_test):
    
    inputsGSR = keras.Input(shape=(vx_trainGSR.shape[1],vx_trainGSR.shape[2]),name="inputGSR")
    inputsPPG = keras.Input(shape=(vx_trainPPG.shape[1],vx_trainPPG.shape[2]),name="inputPPG")    

    x = convolutionalGSR_v(inputsGSR) # weights were freezed
    xppg = convolutionalPPG_v(inputsPPG) # weights were freezed
    y = layers.concatenate([x,xppg])
    xr = layers.Permute((2,1))(y)  #to concatenate one map after another.
    yf = layers.Flatten()(xr)
    d1 = layers.Dropout(rate=0.2)(yf)
    output = layers.Dense(1,activation="sigmoid",kernel_initializer=tf.keras.initializers.GlorotNormal(seed=12345),bias_initializer=initializers.Zeros())(d1)

    valence_model=keras.Model(
                inputs=[inputsGSR,inputsPPG],
                # inputs=[inputsGSR],
                outputs=output, 
                name="valence_NN"
            )    

    valence_model.compile(optimizer=keras.optimizers.Adam(learning_rate=1e-4), loss="binary_crossentropy", metrics=[tf.keras.metrics.BinaryAccuracy()])        

    valence_model.summary()

    history_a = valence_model.fit(
        [vx_trainGSR,vx_trainPPG],  # train model with the whole TRAIN ds. 
        [vy_train],  
        epochs=10,  #6 | Number of epochs to train the model. An epoch is an iteration over the entire x and y data provided.
        # epochs=1,  #Number of epochs to train the model. An epoch is an iteration over the entire x and y data provided.
        batch_size=1,
        shuffle=True, # Boolean (whether to shuffle the training data before each epoch). A kink of evaluation using cv....
    )

    if  (show_plots):        
        plt.plot(history_a.history["loss"], label="Training Loss")
        # plt.plot(history_a.history["val_loss"], label="Validation Loss")
        # plt.plot(history_a.history["val_binary_accuracy"], label="Validation Acc")
        plt.legend()
        plt.show()

    # compare prediction accuracy between validation accuracy (on train ds) and test acc (on TEST ds) 
    # compare training signal vs predicted one 
    # test model on data never seen before:
    prediction = valence_model.predict([vx_testGSR,vx_testPPG])
    # prediction = nn.predict([x_testGSR])

    print("predicted", np.ndarray.flatten(prediction))
    prediction = np.ndarray.round(np.ndarray.flatten(prediction))
    prediction = prediction.astype(int)
    prediction = np.ravel(prediction)
    print("predicted", prediction)
    print("expected", vy_test)

    # val_acc = accuracy_calculation(valence_model,prediction,vy_test)
    # print("accuracy_calculation:", val_acc)

    # val_acc = np.sum(~np.logical_xor(prediction,vy_test))/len(prediction)

    val_acc = accuracy_score(vy_test, prediction)

    print("accuracy:", val_acc)

    [val_cohen, val_uar, val_gm, val_f1] = calc_metrics(vy_test, prediction)

    return [val_cohen, val_uar, val_acc, val_gm, val_f1]


def split_test_alg_DL(gsr_ts, ppg_ts, gsr_label_d, arousal_flag):
    """
    split dataset, then test algorithms, and calculate metrics. 
    arousal_flag: if = 1, then arousal 
    """

    REPETITIONS = 10

    # REPETITIONS = 5

    acc_array = np.zeros(REPETITIONS)
    gmean_array = np.zeros(REPETITIONS)
    f1_score_macro_array = np.zeros(REPETITIONS)
    uar_array = np.zeros(REPETITIONS)
    cohen_array = np.zeros(REPETITIONS)


    for i in range(0,REPETITIONS):
        print("Split Repetition number: ", i)

        [x_trainGSR,x_testGSR, x_trainPPG, x_testPPG, y_train, y_test] = split_ds(gsr_ts, ppg_ts, gsr_label_d)
        convolutionalGSR = train_GSR_AE(x_trainGSR)  
        convolutionalPPG = train_PPG_AE(x_trainPPG)
        if arousal_flag:
            [cohen_array[i], uar_array[i], acc_array[i], gmean_array[i], f1_score_macro_array[i]] = arousal_bimodal_deep_denoising_AE(convolutionalGSR, convolutionalPPG, x_trainGSR, x_testGSR, x_trainPPG, x_testPPG, y_train, y_test) 
        else: 
            [cohen_array[i], uar_array[i], acc_array[i], gmean_array[i], f1_score_macro_array[i]] = valence_bimodal_deep_denoising_AE(convolutionalGSR, convolutionalPPG, x_trainGSR, x_testGSR, x_trainPPG, x_testPPG, y_train, y_test) 
        print("-------- Model Performance ----------: ")    
        print("accuracy: ", acc_array)
        print("gmean: ", gmean_array)
        print("f1_score: ", f1_score_macro_array)    
        print("UAR: ", uar_array)
        print("Cohen Kappa score: ", cohen_array)

    # return np.array([np.nanmean(acc_array),np.nanmean(gmean_array),np.nanmean(f1_score_macro_array)],dtype='float')
    return np.array([np.nanmean(cohen_array), np.nanmean(uar_array), np.nanmean(acc_array),np.nanmean(gmean_array),np.nanmean(f1_score_macro_array)],dtype='float')

def evaluate_alg_DL(gsr_ts_arousal, ppg_ts_arousal, gsr_arousal_d, gsr_ts_valence, ppg_ts_valence, gsr_valence_d):
    #evaluate model (arousal and valence separatedly): accuraccy, geometric mean and macro f1 score
    
    [aro_cohen, aro_uar, aro_acc, aro_gm, aro_f1] = split_test_alg_DL(gsr_ts_arousal, ppg_ts_arousal, gsr_arousal_d, 1)
    [val_cohen, val_uar, val_acc, val_gm, val_f1] = split_test_alg_DL(gsr_ts_valence, ppg_ts_valence, gsr_valence_d, 0)

    return np.array([(val_cohen, val_uar, val_acc,val_gm, val_f1, aro_cohen, aro_uar,aro_acc,aro_gm, aro_f1)])


def evaluate_alg_DL_arousal(gsr_ts_arousal, ppg_ts_arousal, gsr_arousal_d, gsr_ts_valence, ppg_ts_valence, gsr_valence_d):
    #evaluate model (arousal only): accuraccy, geometric mean and macro f1 score
    
    [aro_cohen, aro_uar, aro_acc, aro_gm, aro_f1] = split_test_alg_DL(gsr_ts_arousal, ppg_ts_arousal, gsr_arousal_d, 1)
    # [val_cohen, val_uar, val_acc, val_gm, val_f1] = split_test_alg_DL(gsr_ts_valence, ppg_ts_valence, gsr_valence_d, 0)

    return np.array([(np.nan, np.nan, np.nan,np.nan, np.nan, aro_cohen, aro_uar,aro_acc,aro_gm, aro_f1)]) #valence metris = np.nan

def evaluate_alg_DL_valence(gsr_ts_arousal, ppg_ts_arousal, gsr_arousal_d, gsr_ts_valence, ppg_ts_valence, gsr_valence_d):
    #evaluate model (valence only): accuraccy, geometric mean and macro f1 score
    
    # [aro_cohen, aro_uar, aro_acc, aro_gm, aro_f1] = split_test_alg_DL(gsr_ts_arousal, ppg_ts_arousal, gsr_arousal_d, 1)
    [val_cohen, val_uar, val_acc, val_gm, val_f1] = split_test_alg_DL(gsr_ts_valence, ppg_ts_valence, gsr_valence_d, 0)

    return np.array([(val_cohen, val_uar, val_acc,val_gm, val_f1, np.nan, np.nan,np.nan,np.nan, np.nan)]) #arousal metris = np.nan


def array_is_not_empty(array):
    
    flag = np.size(array)
    if flag != 0:
        print("Array is not empty")
        flag = 1
    else:
        print("Array is empty")
        flag = 0 
    return flag


def delete_medium_labels(fv, arousal, valence):
    # Build feature vector (and corresponding label array) without medium labels
    # new feature vector: same as the original feature vector but with deleted
    # rows (see valence_medium_indexes and arousal_medium_indexes)
    # returns:
    # new feature vectors: one for arousal (with its new corresponding labels, without medium labels) and other for valence
    # [fv_arousal, fv_valence, arousal, valence]

    # Find "medium" partition indices (values from 4 to 7): 
    arousal_medium_indexes = []

    for i in range(np.shape(arousal)[0]):
        if arousal[i] == [MEDIUM_LABEL_VALUES]:
          arousal_medium_indexes.append(i)

    print("indexes of the arousal array that have medium labels: ")
    print(arousal_medium_indexes) # indexes of the arousal array that have medium labels

    valence_medium_indexes = []

    for i in range(np.shape(valence)[0]):
        if valence[i] == [MEDIUM_LABEL_VALUES]:
          valence_medium_indexes.append(i)

    print("indexes of the valence array that have medium labels: ")
    print(valence_medium_indexes)  # indexes of the valence array that have medium labels

    # if medium partition is to be deleted, then indexes are important to know.

    fv_arousal = fv.copy()

    if array_is_not_empty(arousal_medium_indexes):
        # arousalbck = arousal.copy()
        
        for row in arousal_medium_indexes:
            arousal_test = np.delete(arousal,arousal_medium_indexes,0) 
            fv_arousal_test = np.delete(fv_arousal,arousal_medium_indexes,0) 

        fv_arousal = fv_arousal_test.copy()
        arousal = arousal_test.copy()
        del fv_arousal_test
        del arousal_test    
    else: 
        fv_arousal = fv.copy()

    fv_valence = fv.copy()

    if array_is_not_empty(valence_medium_indexes):
        # valencebck = valence.copy()
        
        for row in valence_medium_indexes:
            valence_test = np.delete(valence,valence_medium_indexes,0) 
            fv_valence_test = np.delete(fv_valence,valence_medium_indexes,0) 

        fv_valence = fv_valence_test.copy()
        valence = valence_test.copy()
        del fv_valence_test
        del valence_test
    else:
        fv_valence = fv.copy()

    print("number of AROUSAL instances (after deletion of medium indexes or Nan/masked elements): ")  # added Nan/masked elements because exception_trials management is done before this function was called
    print(np.max(np.shape(arousal)))
    print("number of VALENCE instances (after deletion of medium indexes or Nan/masked elements): ")
    print(np.max(np.shape(valence)))

    return [fv_arousal, fv_valence, arousal, valence]


def isnan_or_masked(features):
    """
    Return True if numpy array "features" has a NaN or masked element.  
    """
    has_nan_or_masked = False
    features = np.reshape(features,(1, np.max(np.shape(features))))
    
    for index in range(np.shape(features)[1]):
        if any(isnan(x) for x in features[:,index]) or any(np.ma.is_masked(x) for x in features[:,index]):
            has_nan_or_masked = True
    return has_nan_or_masked


def delete_nanmasked_labels(arousal, valence, indexes):
    """
    Delete labels associated to discarted trials (because of NaN o masked element). 
    Delete all arousal and valence rows indicated by "indexes" array.
    """    
    if array_is_not_empty(indexes):        
        for row in indexes:
            arousal_test = np.delete(arousal,indexes,0) 
            valence_test = np.delete(valence,indexes,0) 
        arousal = arousal_test.copy()
        valence = valence_test.copy()
    
    return [arousal, valence]


if __name__ == "__main__":

        # # Record the start time
        # start_time = time.time()

        # print(map_labels(9))
        # base as a function of the platform:
        if platform == "linux" or platform == "linux2":

            # current_folder=$(pwd)
            # file_path="$current_folder/subfolder_name/file_name"
            physiological_base = r"./physiological/sub_"
            annotation_base = r"./annotations/sub_"
            seqs_order_num_base = r"seqs_"
        elif platform == "win32":
            physiological_base = r"C:\Users\mbamo\Desktop\Datasets\CASE\interpolated\physiological\sub_"
            annotation_base = r"C:\Users\mbamo\Desktop\Datasets\CASE\interpolated\annotations\sub_"
            seqs_order_num_base = r"C:\Users\mbamo\Desktop\Datasets\CASE\metadata\seqs_"

        # Specify the path to your text file
        participant = '28'
        physiological_file = physiological_base + participant + '.csv'
        annotation_file = annotation_base + participant + '.csv'
        seqs_order_num_file = seqs_order_num_base + 'order_num_csv' + '.csv'

        # Use numpy's loadtxt function to load the data
        data = pd.read_csv(physiological_file)
        seqs_order_num = pd.read_csv(seqs_order_num_file)

        # time_vector = data['daqtime']/(1000*60)  # [minutes] 
        time_vector = data['daqtime']  # [ms] 

        # load annotations
        annotations = pd.read_csv(annotation_file)
        valence = annotations['valence']
        arousal = annotations['arousal']

        # plt.plot(valence)
        # plt.show(block=True)

        stimulus_tag_list = [map_video_to_tag(i) for i in data['video']] 

        data['tag'] = stimulus_tag_list  # dataframe now contains the stimulus tag of each video

        [ gsr_mc, ppg_mc ] = baseline_mean_centered_normalization(data, show_plot=0)

        # plot_signals_with_stimulus_timestamps(ppg_mc,gsr_mc, data['tag'])

        gsr_signals, gsr_info = filter_gsr(gsr_mc, show_plot=0)

        ppg_signals = filter_ppg(ppg_mc, show_plot=0)

        # compare_onsets(gsr_signals, gsr_info ,data['tag'])

        # test windowing
        window_size = 20*1000  # ms
        step = 9*1000 #ms
        overlap = False
        
        combined_df = perform_windowing(data, ppg_signals, gsr_signals, valence, arousal, seqs_order_num, participant, window_size, step, overlap, show_plot = 0)

        # verify_performed_windowing(combined_df, gsr_signals)

        print("Unified DataFrame Shape: ", combined_df.shape)

        combined_df_median = median_voting(combined_df)

        # verify_median_voting(combined_df, combined_df_median)
 
        # # combined_df_mapped = map_valence_arousal(combined_df_median, scheme = 'tripartite', number_thresholds = 2, threshold = 5, L = 3.6, H = 6.4)
        # combined_df_mapped = map_valence_arousal(combined_df_median, scheme = 'bipartite', number_thresholds = 1, threshold = 5, L = 3.6, H = 4.4)

        # # check_correct_mapping(combined_df, combined_df_median, combined_df_mapped)

        # X_windowed, y_median = build_dataset(combined_df_mapped)

        # xf = feature_extract_GSR_PPG_non_linear(X_windowed)

        # # check_poincare_plots(X_windowed, show = True)

        # performance = evaluate_alg_CASE(GradientBoostingClassifier(n_estimators=15),xf,y_median, report=True)
        # performance
        # performance = evaluate_alg_CASE(SVC(C=0.1),xf,y_median, report=True)
        # performance
        # # performance = evaluate_alg_CASE(SVC(C=0.1),xf,y_median, report=True)
        # performance

        # # # Record the end time
        # # end_time = time.time()

        # # # Calculate the total execution time
        # # execution_time = end_time - start_time

        # # # Print the execution time
        # # print("Total execution time:", execution_time, "seconds")

