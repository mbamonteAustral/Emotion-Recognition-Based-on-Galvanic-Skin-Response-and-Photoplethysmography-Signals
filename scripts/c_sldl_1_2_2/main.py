"""
main.py

main script. It calls processDEAPdata.py to apply different ALG to 
each participant data.
"""

#%% --------------------------------------------------
from processdataf import process_data
import numpy as np
from numpy import savetxt
from sys import platform
from openpyxl import load_workbook
import time
import tensorflow as tf

with tf.device('/CPU:0'):

    # Record the start time
    start_time = time.time()

    participants = np.arange(1, 31)  # participant from 1 to 30

    # participants = np.arange(1, 3)  # participant from 1 to 30

    # splitterseed = 12345
    # estimatorseed = 12345

    index = 1 # usefull if first participant is not the number 1. Index ensure a first participant performance is loaded, independentedly participant number is 1 or any other.
    # if index = 1,  then a first participant is NOT loaded yet.

    #%% --------------------------------------------------
    # check if algorithm performance files existes:  
    import os.path

    alg_files_created = 0    # = 1 is algorithm performance files are already present  

    knn_file_exists = os.path.exists('knn_perm.csv')
    dt_file_exists = os.path.exists('dt_perm.csv')
    rf_file_exists = os.path.exists('rf_perm.csv')
    svm_file_exists = os.path.exists('svm_perm.csv')
    gbm_file_exists = os.path.exists('gbm_perm.csv')
    dummy_perm_exists = os.path.exists('dummy_perm.csv')

    # if knn_file_exists and dt_file_exists and rf_file_exists and svm_file_exists and gbm_file_exists and dummy_perm_exists:
    #     print("all algorithm performance files were created correctly before")
    #     alg_files_created = 1

    #     from numpy import loadtxt
    #     print("loading data already saved in algorithm performance files")
    #     # load array
    #     knn_perm = loadtxt('knn_perm.csv', delimiter=',')
    #     print(knn_perm)
        
    #     dt_perm = loadtxt('dt_perm.csv', delimiter=',')
    #     print(dt_perm)
        
    #     rf_perm = loadtxt('rf_perm.csv', delimiter=',')
    #     print(rf_perm)
        
    #     svm_perm = loadtxt('svm_perm.csv', delimiter=',')
    #     print(svm_perm)
        
    #     gbm_perm = loadtxt('gbm_perm.csv', delimiter=',')
    #     print(gbm_perm)

    #     dummy_perm = loadtxt('dummy_perm.csv', delimiter=',')
    #     print(dummy_perm)

            
    for i in participants:
        # if platform == "linux" or platform == "linux2":
        #     participant = str(i)
        #     fname = r"E4_"
        #     fname_label = r"P"
        #     ppg_workbook = load_workbook(filename=fname + r"BVP_p" + participant + r"t.xlsx", read_only=True, data_only=True)
        #     gsr_workbook = load_workbook(filename=fname + r"EDA_p" + participant + r"t.xlsx", read_only=True, data_only=True)
        #     labels_workbook = load_workbook(filename=fname_label + participant + r".self.xlsx", read_only=True, data_only=True)
            
        # elif platform == "win32":
        #     participant = str(i)
        #     fname = r"C:\Users\mbamo\Desktop\Datasets\K-EmoCon\data for vpn pc\E4_"
        #     fname_label = r"C:\Users\mbamo\Desktop\Datasets\K-EmoCon\data for vpn pc\P"
        #     ppg_workbook = load_workbook(filename=fname + r"BVP_p" + participant + r"t.xlsx", read_only=True, data_only=True)
        #     gsr_workbook = load_workbook(filename=fname + r"EDA_p" + participant + r"t.xlsx", read_only=True, data_only=True)
        #     labels_workbook = load_workbook(filename=fname_label + participant + r".self.xlsx", read_only=True, data_only=True)

        
        print("    ")
        print("    ")
        print("-------------------------------")
        print("participant: ", i)
        # print(participant)
        print("-------------------------------")
        print("    ")
        print("    ")
        try:
            performance, window_size, step, overlap, instances = process_data(str(i))
            # performance = process_data(participant, splitterseed, estimatorseed)
                
            # performance = np.array([[0.75      , 0.84375   , 0.5       , 0.56666667],
            # [0.625     , 0.625     , 0.625     , 0.63333333],
            # [0.625     , 0.75      , 0.5       , 0.53333333],
            # [0.625     , 0.375     , 0.625     , 0.6       ],
            # [0.625     , 0.875     , 0.5       , 0.53333333]])

            if (alg_files_created == 0):     #load first participant performance    
                knn_perm = performance[0]
                dt_perm = performance[1]
                rf_perm = performance[2]
                svm_perm = performance[3]
                gbm_perm = performance[4]
                bddae_perm = performance[5]
                dummy_perm = performance[6]

                # index = index + 1 
                alg_files_created = 1
                print("first participant performance loaded")
            else:
                knn_perm = np.vstack([knn_perm,performance[0]])
                dt_perm = np.vstack([dt_perm,performance[1]])
                rf_perm = np.vstack([rf_perm,performance[2]])
                svm_perm = np.vstack([svm_perm,performance[3]])
                gbm_perm = np.vstack([gbm_perm,performance[4]])   
                bddae_perm = np.vstack([bddae_perm,performance[5]])   
                dummy_perm = np.vstack([dummy_perm,performance[6]])     
                print("participant performance loaded")

            # savetxt('knn_perm.csv',knn_perm, delimiter=',')
            # savetxt('dt_perm.csv',dt_perm, delimiter=',')
            # savetxt('rf_perm.csv',rf_perm, delimiter=',')
            # savetxt('svm_perm.csv',svm_perm, delimiter=',')
            # savetxt('gbm_perm.csv',gbm_perm, delimiter=',')
            # savetxt('dummy_perm.csv',dummy_perm, delimiter=',')
    
        except Exception as exc:
            print("-----------Exception ----------------")
            print(exc)
            print("-------------------------------------")

            
        print("    ")
        print("    ")
        print("-------------------------------")
        print("LAST participant's saved data: ")
        print(i)
        print("-------------------------------")
        print("    ")
        print("    ")
        

    #%% -------------------------
    # print algorithms performances:

    print("---------------------------")
    print("KNN performance:")
    print(knn_perm)
    print("KNN mean:")
    print(np.nanmean(knn_perm,axis=0))

    print("---------------------------")

    print("---------------------------")
    print("DT performance:")
    print(dt_perm)
    print("DT mean:")
    print(np.nanmean(dt_perm,axis=0))

    print("---------------------------")

    print("---------------------------")
    print("RF performance:")
    print(rf_perm)
    print("RF mean:")
    print(np.nanmean(rf_perm,axis=0))

    print("---------------------------")

    print("---------------------------")
    print("SVM performance:")
    print(svm_perm)
    print("SVM mean:")
    print(np.nanmean(svm_perm,axis=0))

    print("---------------------------")

    print("---------------------------")
    print("GBM performance:")
    print(gbm_perm)
    print("GBM mean:")
    print(np.nanmean(gbm_perm,axis=0))

    print("---------------------------")

    print("---------------------------")
    print("BDDAE performance:")
    print(bddae_perm)
    print("BDDAE mean:")
    print(np.nanmean(bddae_perm,axis=0))

    print("---------------------------")

    print("---------------------------")
    print("DUMMY performance:")
    print(dummy_perm)
    print("DUMMY mean:")
    print(np.nanmean(dummy_perm,axis=0))

    print("---------------------------")

    #%%----------------------------------------------------------------
    # build final performace vector
    performance_final = np.array([(0.0, 0.0,0.0, 0.0, 0.0, 0.0,0.0,0.0,0.0,0.0)]) # store ALG metrics
    performance_final[-1] = np.nanmean(knn_perm,axis=0)
    performance_final = np.vstack([performance_final,np.nanmean(dt_perm,axis=0)])
    performance_final = np.vstack([performance_final,np.nanmean(rf_perm,axis=0)])
    performance_final = np.vstack([performance_final,np.nanmean(svm_perm,axis=0)])
    performance_final = np.vstack([performance_final,np.nanmean(gbm_perm,axis=0)])
    performance_final = np.vstack([performance_final,np.nanmean(bddae_perm,axis=0)])
    performance_final = np.vstack([performance_final,np.nanmean(dummy_perm,axis=0)])

    import os
    print("Current folder:", os.getcwd())
    print("-------------------------------------")
    print("----- RESULTS ------")
    print("-------------------------------------")
    print("Window size (sec): ", window_size/1000)
    print("step (sec): ", step/1000)
    print("overlap: ", overlap)
    if overlap:
        print("perc. of overlap: ", (window_size-step)*100/window_size)
        print("overlap duration (sec): ", (window_size-step)/1000)
    print("Number of windows / instances: ", instances)
    print("---------------------------")
    print("ALG Means: ")
    print("-rows: alg : KNN = 0; DT = 1; RF = 2; SVM = 3; GBM = 4; BDDAE = 5; DUMMY = 5")
    print("columns:")
    print("'val_cohen','val_uar', 'val_acc', 'val_gm',  'val_f1',  'aro_cohen','aro_uar', 'aro_acc', 'aro_gm',  'aro_f1'")
    print("-------------------------------------------------------------")
    print("  v_c   v_u   v_a   v_g   v_f1  a_c   a_u   a_a   a_g   a_f1")
    print(np.round(performance_final, decimals=3))


    #%% -------------------------
    # print("---------------------------")
    # print("ALG Means: ")
    # print("1) KNN, 2) DT, 3) RF, 4) SVM, 5) GBM; 6) BDDAE, 7) DUMMY")
    # print("columns: 'val_cohen', 'val_uar', 'val_acc','val_gm', 'val_f1', 'aro_cohen', 'aro_uar','aro_acc','aro_gm', 'aro_f1' ")  

    # print(np.nanmean(knn_perm,axis=0))
    # print(np.nanmean(dt_perm,axis=0))
    # print(np.nanmean(rf_perm,axis=0))
    # print(np.nanmean(svm_perm,axis=0))
    # print(np.nanmean(gbm_perm,axis=0))
    # print(np.nanmean(bddae_perm,axis=0))
    # print(np.nanmean(dummy_perm,axis=0))

    # print("---------------------------")

    # %%
    # ---- Extract ALG metrics ---- 
    # alg : knn = 0; dt = 1; rf = 2; svm = 3; gbm = 4

    # %% -----------------------------------
    # save performance in .csv file:

    # from numpy import savetxt
    savetxt('knn_perm.csv',knn_perm, delimiter=',')
    savetxt('dt_perm.csv',dt_perm, delimiter=',')
    savetxt('rf_perm.csv',rf_perm, delimiter=',')   
    savetxt('svm_perm.csv',svm_perm, delimiter=',')
    savetxt('gbm_perm.csv',gbm_perm, delimiter=',')
    savetxt('bddda_perm.csv',bddae_perm, delimiter=',')
    savetxt('dummy_perm.csv',dummy_perm, delimiter=',')


    # Record the end time
    end_time = time.time()

    # Calculate the elapsed time
    elapsed_time = end_time - start_time  # in seconds

    print("Elapsed time:", elapsed_time/(60), "minutes")
    print("Elapsed time:", elapsed_time/(60*60), "hours")

    #%% -------------------------
    # ----------------------------------------------------------------
    # ----------------------------------------------------------------
    # ----------------------------------------------------------------
    # ----------------------------------------------------------------
    # ----------------------------------------------------------------
    # ----------------------------------------------------------------
    # ----------------------------------------------------------------
    # ----------------------------------------------------------------




    #%% ----------------------------------------
    # # load numpy array from csv file

    # from numpy import loadtxt

    # # # load array
    # knn_perm = loadtxt('knn_perm.csv', delimiter=',')
    # print(knn_perm)

    # dt_perm = loadtxt('dt_perm.csv', delimiter=',')
    # print(dt_perm)

    # rf_perm = loadtxt('rf_perm.csv', delimiter=',')
    # print(rf_perm)

    # svm_perm = loadtxt('svm_perm.csv', delimiter=',')
    # print(svm_perm)

    # gbm_perm = loadtxt('gbm_perm.csv', delimiter=',')
    # print(gbm_perm)


    # #%% -------------------------
    # # print algorithms performances:

    # print("---------------------------")
    # print("KNN performance:")
    # print(knn_perm)
    # print("KNN mean:")
    # print(knn_perm.mean(axis=0))
    # print("---------------------------")

    # print("---------------------------")
    # print("DT performance:")
    # print(dt_perm)
    # print("DT mean:")
    # print(dt_perm.mean(axis=0))
    # print("---------------------------")

    # print("---------------------------")
    # print("RF performance:")
    # print(rf_perm)
    # print("RF mean:")
    # print(rf_perm.mean(axis=0))
    # print("---------------------------")

    # print("---------------------------")
    # print("SVM performance:")
    # print(svm_perm)
    # print("SVM mean:")
    # print(svm_perm.mean(axis=0))
    # print("---------------------------")

    # print("---------------------------")
    # print("GBM performance:")
    # print(gbm_perm)
    # print("GBM mean:")
    # print(gbm_perm.mean(axis=0))
    # print("---------------------------")

    # #%% -------------------------
    # print("---------------------------")
    # print("ALG Means: ")
    # print("1) KNN, 2) DT, 3) RF, 4) SVM, 5) GBM, 6) DUMMY")
    # print(knn_perm.mean(axis=0))
    # print(dt_perm.mean(axis=0))
    # print(rf_perm.mean(axis=0))
    # print(svm_perm.mean(axis=0))
    # print(gbm_perm.mean(axis=0))
    # print("---------------------------")


    # #%% ----------------------------------------
    # # load numpy array from csv file

    # from numpy import loadtxt
    # # load array
    # knndata = loadtxt('knn_perm.csv', delimiter=',')
    # print(knndata)

    # dtdata = loadtxt('dt_perm.csv', delimiter=',')
    # print(dtdata)

    # rfdata = loadtxt('rf_perm.csv', delimiter=',')
    # print(rfdata)

    # svmdata = loadtxt('svm_perm.csv', delimiter=',')
    # print(svmdata)

    # gbmdata = loadtxt('gbm_perm.csv', delimiter=',')
    # print(gbmdata)

    # #%%----------------------------------------------------------------
    # # --- box plot - labels=('val_acc', 'aro_acc')
    # # ----------------------------------------------------------------
    # import matplotlib.pyplot as plt
    # import numpy as np

    # valaro=np.append(knndata[:,0:1],knndata[:,2:3],axis=1)
    # plt.boxplot(valaro,labels=('val_acc', 'aro_acc'))
    # plt.title('KNN performance')
    # plt.show()

    # plt.boxplot(knndata,labels=('val_acc', 'val_auc','aro_acc','aro_auc'))
    # plt.title('KNN performance')
    # plt.show()



    # #%%----------------------------------------------------------------
    # # --- box plot - labels=('val_acc', 'aro_acc')
    # # ----------------------------------------------------------------

    # valaro=np.append(dtdata[:,0:1],dtdata[:,2:3],axis=1)
    # plt.boxplot(valaro,labels=('val_acc', 'aro_acc'))
    # plt.title('DT performance')
    # plt.show()

    # plt.boxplot(dtdata,labels=('val_acc', 'val_auc','aro_acc','aro_auc'))
    # plt.title('DT performance')
    # plt.show()


    # #%%----------------------------------------------------------------
    # # --- box plot - labels=('val_acc', 'aro_acc')
    # # ----------------------------------------------------------------

    # valaro=np.append(rfdata[:,0:1],rfdata[:,2:3],axis=1)
    # plt.boxplot(valaro,labels=('val_acc', 'aro_acc'))
    # plt.title('RF performance')
    # plt.show()

    # plt.boxplot(rfdata,labels=('val_acc', 'val_auc','aro_acc','aro_auc'))
    # plt.title('RF performance')
    # plt.show()



    # #%%----------------------------------------------------------------
    # # --- box plot - labels=('val_acc', 'aro_acc')
    # # ----------------------------------------------------------------

    # valaro=np.append(svmdata[:,0:1],svmdata[:,2:3],axis=1)
    # plt.boxplot(valaro,labels=('val_acc', 'aro_acc'))
    # plt.title('SVM performance')
    # plt.show()

    # plt.boxplot(svmdata,labels=('val_acc', 'val_auc','aro_acc','aro_auc'))
    # plt.title('SVM performance')
    # plt.show()



    # #%%----------------------------------------------------------------
    # # --- box plot - labels=('val_acc', 'aro_acc')
    # # ----------------------------------------------------------------

    # valaro=np.append(gbmdata[:,0:1],gbmdata[:,2:3],axis=1)
    # plt.boxplot(valaro,labels=('val_acc', 'aro_acc'))
    # plt.title('GBM performance')
    # plt.show()

    # plt.boxplot(gbmdata,labels=('val_acc', 'val_auc','aro_acc','aro_auc'))
    # plt.title('GBM performance')
    # plt.show()


    # # #%%----------------------------------------------------------------
    # # # --- box plot - labels=('val_acc', 'val_auc','aro_acc','aro_auc')

    # # import matplotlib.pyplot as plt

    # # # plt.boxplot(knndata[:,1:5],labels=('val_acc', 'val_auc','aro_acc','aro_auc'))
    # # plt.boxplot(knndata,labels=('val_acc', 'val_auc','aro_acc','aro_auc'))
    # # plt.title('KNN performance')
    # # plt.show()


    # # #%%----------------------------------------------------------------
    # # # --- box plot - labels=('val_acc', 'val_auc','aro_acc','aro_auc')

    # # # plt.boxplot(dtdata[:,1:5],labels=('val_acc', 'val_auc','aro_acc','aro_auc'))
    # # plt.boxplot(dtdata,labels=('val_acc', 'val_auc','aro_acc','aro_auc'))
    # # plt.title('DT performance')
    # # plt.show()

    # # #%%----------------------------------------------------------------
    # # # --- box plot - labels=('val_acc', 'val_auc','aro_acc','aro_auc')

    # # plt.boxplot(rfdata,labels=('val_acc', 'val_auc','aro_acc','aro_auc'))
    # # plt.title('RF performance')
    # # plt.show()


    # # #%%----------------------------------------------------------------
    # # # --- box plot - labels=('val_acc', 'val_auc','aro_acc','aro_auc')

    # # plt.boxplot(svmdata,labels=('val_acc', 'val_auc','aro_acc','aro_auc'))
    # # plt.title('svm performance')
    # # plt.show()


    # # #%%----------------------------------------------------------------
    # # # --- box plot - labels=('val_acc', 'val_auc','aro_acc','aro_auc')

    # # plt.boxplot(gbmdata,labels=('val_acc', 'val_auc','aro_acc','aro_auc'))
    # # plt.title('gbm performance')
    # # plt.show()

    # #%%----------------------------------------------------------------
    # # --- ALG performance per participant

    # participants = [1,2,4,5,6,7,8,9,10,11,12,15,19,16,20,21]  #16 participants

    # for i in range(0,len(participants)):
    #     print("participant n.", participants[i])
    #     print("     val_acc -- val_auc -- aro_acc --- aro_auc")
    #     print("KNN", knndata[i,:])
    #     print("DT",dtdata[i,:])
    #     print("RF", rfdata[i,:])  
    #     print("SVM", svmdata[i,:])
    #     print("GBM", gbmdata[i,:])
    #     print("")



    # participants = [1,2,4,5,6,7,8,9,10,11,12,15,19,16,20,21]  #16 participants

    # for i in range(0,len(participants)):
    #     print("participant n.", participants[i])
    #     print("     val_acc -- val_auc -- aro_acc --- aro_auc")
    #     print("KNN", knndata[i,:])
    #     print("DT",dtdata[i,:])
    #     print("RF", rfdata[i,:])  
    #     print("SVM", svmdata[i,:])
    #     print("GBM", gbmdata[i,:])
    #     print("")

